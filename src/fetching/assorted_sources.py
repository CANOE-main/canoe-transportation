"""Fetch, validate, and source-normalize small public transportation inputs."""

import argparse
import io
import json
import logging
import math
import os
import re
from dataclasses import dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Literal
from urllib.parse import urljoin, urlparse
from zipfile import BadZipFile, ZipFile

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import (
    column_index_from_string,
    range_boundaries,
)
from pydantic import BaseModel, ConfigDict, model_validator
from pypdf import PdfReader
from pypdf.errors import PdfReadError

from utils import (
    ConfigBundle,
    file_sha256,
    load_config_bundle,
    load_harmonization_rules,
    resolve_input_path,
)
from validation.config_models import SourceSpec


class AssortedSourcesError(ValueError):
    """Raised when a configured small-source contract is missing or has changed."""


class ArtifactRequest(BaseModel):
    """Resolved immutable acquisition and cache contract for one physical artifact."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    source_id: str
    component_id: str
    source_version: str
    url: str
    cache_path: Path
    file_type: Literal["zip", "xlsx", "csv", "html", "javascript", "pdf"]
    expected_sha256: str
    expected_bytes: int

    @model_validator(mode="after")
    def validate_request(self) -> "ArtifactRequest":
        parsed = urlparse(self.url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError(f"Artifact URL is invalid: {self.url}")
        if not self.cache_path.is_absolute():
            raise ValueError(f"Cache path must be absolute: {self.cache_path}")
        expected_suffixes = {
            "zip": {".zip"},
            "xlsx": {".xlsx"},
            "csv": {".csv"},
            "html": {".html", ".htm"},
            "javascript": {".js"},
            "pdf": {".pdf"},
        }
        if self.cache_path.suffix.casefold() not in expected_suffixes[self.file_type]:
            raise ValueError(
                f"{self.file_type} cache has an invalid suffix: {self.cache_path}"
            )
        if not re.fullmatch(r"[0-9a-f]{64}", self.expected_sha256):
            raise ValueError("Expected SHA-256 must contain 64 lowercase hex characters")
        if self.expected_bytes <= 0:
            raise ValueError("Expected byte count must be positive")
        return self


@dataclass(frozen=True)
class SelectionStats:
    """Auditable source-record selection counts."""

    input_records: int
    selected_records: int
    excluded_records: int


MANIFEST_COLUMNS = [
    "source_id",
    "component_id",
    "source_version",
    "acquisition_url",
    "cached_file",
    "sha256",
    "bytes",
    "cache_status",
    "input_records",
    "selected_records",
    "excluded_records",
    "selection",
    "output_files",
    "warning_count",
    "status",
]


def module_rules(bundle: ConfigBundle) -> dict[str, Any]:
    """Load assorted-source extraction and output rules."""
    return load_harmonization_rules(bundle, "assorted_sources")


def _source(bundle: ConfigBundle, source_id: str) -> SourceSpec:
    source = bundle.sources.sources.get(source_id)
    if not isinstance(source, SourceSpec):
        raise AssortedSourcesError(f"sources.yaml missing {source_id}")
    return source


def _top_level_request(
    bundle: ConfigBundle,
    *,
    source_id: str,
    component_id: str,
    file_type: Literal["zip", "xlsx", "csv", "html"],
) -> ArtifactRequest:
    source = _source(bundle, source_id)
    adapter = source.adapter
    access = adapter.get("access", {})
    return ArtifactRequest(
        source_id=source_id,
        component_id=component_id,
        source_version=source.version,
        url=str(access.get("url", "")),
        cache_path=resolve_input_path(bundle, "cache", str(adapter.get("cache_path", ""))),
        file_type=file_type,
        expected_sha256=str(adapter.get("expected_sha256", "")),
        expected_bytes=int(adapter.get("expected_bytes", 0)),
    )


def _component_request(
    bundle: ConfigBundle,
    *,
    source_id: str,
    component_id: str,
    file_type: Literal["pdf"],
) -> ArtifactRequest:
    source = _source(bundle, source_id)
    component = source.component(component_id)
    adapter = component.adapter
    return ArtifactRequest(
        source_id=source_id,
        component_id=component_id,
        source_version=component.version or source.version,
        url=str(adapter.get("url", "")),
        cache_path=resolve_input_path(bundle, "cache", str(adapter.get("cache_path", ""))),
        file_type=file_type,
        expected_sha256=str(adapter.get("expected_sha256", "")),
        expected_bytes=int(adapter.get("expected_bytes", 0)),
    )


def build_requests(
    bundle: ConfigBundle, rules: dict[str, Any] | None = None
) -> dict[str, ArtifactRequest]:
    """Resolve the six directly addressed source artifacts from configuration."""
    selected = rules or module_rules(bundle)
    nhtsa = selected["nhtsa_cafe"]
    nems = selected["eia_nems"]
    gcam = selected["jgcri_gcam"]
    regen = selected["epri_us_regen"]
    faa = selected["faa"]
    return {
        "nhtsa": _top_level_request(
            bundle,
            source_id=str(nhtsa["source_id"]),
            component_id=str(nhtsa["component_id"]),
            file_type="zip",
        ),
        "nems": _top_level_request(
            bundle,
            source_id=str(nems["source_id"]),
            component_id=str(nems["component_id"]),
            file_type="xlsx",
        ),
        "gcam": _top_level_request(
            bundle,
            source_id=str(gcam["source_id"]),
            component_id=str(gcam["component_id"]),
            file_type="csv",
        ),
        "regen_page": _top_level_request(
            bundle,
            source_id=str(regen["source_id"]),
            component_id=str(regen["component_id"]),
            file_type="html",
        ),
        "faa_section_3": _component_request(
            bundle,
            source_id=str(faa["source_id"]),
            component_id=str(faa["capacity_component_id"]),
            file_type="pdf",
        ),
        "faa_section_4": _component_request(
            bundle,
            source_id=str(faa["source_id"]),
            component_id=str(faa["costs_component_id"]),
            file_type="pdf",
        ),
    }


def fetch_to_cache(
    request: ArtifactRequest,
    *,
    session: requests.Session | None = None,
    timeout: int = 120,
) -> str:
    """Atomically download one configured artifact, or reuse its existing cache."""
    if request.cache_path.is_file():
        return "cached"
    request.cache_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = request.cache_path.with_suffix(request.cache_path.suffix + ".part")
    client = session or requests.Session()
    try:
        response = client.get(request.url, stream=True, timeout=timeout)
        response.raise_for_status()
        with temporary.open("wb") as handle:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    handle.write(chunk)
        if not temporary.is_file() or temporary.stat().st_size == 0:
            raise AssortedSourcesError(f"Downloaded artifact is empty: {request.url}")
        _validate_artifact_path(request, temporary)
        os.replace(temporary, request.cache_path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return "downloaded"


def ensure_cached(
    request: ArtifactRequest,
    *,
    download: bool,
    session: requests.Session | None = None,
) -> str:
    """Enforce declared online/offline cache behavior."""
    if download:
        return fetch_to_cache(request, session=session)
    if not request.cache_path.is_file():
        raise FileNotFoundError(
            f"Offline cache missing for {request.source_id}/{request.component_id}: "
            f"{request.cache_path}"
        )
    return "cached"


def _validate_artifact_path(request: ArtifactRequest, path: Path) -> str:
    actual_bytes = path.stat().st_size
    if actual_bytes != request.expected_bytes:
        raise AssortedSourcesError(
            f"{request.source_id}/{request.component_id} byte count changed: "
            f"expected {request.expected_bytes}, got {actual_bytes}"
        )
    actual_hash = file_sha256(path)
    if actual_hash != request.expected_sha256:
        raise AssortedSourcesError(
            f"{request.source_id}/{request.component_id} SHA-256 changed: "
            f"expected {request.expected_sha256}, got {actual_hash}"
        )
    with path.open("rb") as handle:
        prefix = handle.read(16)
    if request.file_type in {"zip", "xlsx"} and not prefix.startswith(b"PK"):
        raise AssortedSourcesError(f"{path} is not an OOXML/ZIP artifact")
    if request.file_type == "pdf" and not prefix.startswith(b"%PDF-"):
        raise AssortedSourcesError(f"{path} is not a PDF artifact")
    if request.file_type == "html" and b"<" not in prefix:
        raise AssortedSourcesError(f"{path} is not an HTML artifact")
    if request.file_type == "csv" and prefix.lstrip().startswith(b"<"):
        raise AssortedSourcesError(f"{path} is HTML, not CSV")
    return actual_hash


def validate_cached_identity(request: ArtifactRequest) -> str:
    """Validate configured bytes, hash, and coarse physical artifact type."""
    if not request.cache_path.is_file():
        raise FileNotFoundError(request.cache_path)
    return _validate_artifact_path(request, request.cache_path)


def _numeric(
    value: object,
    *,
    context: str,
    minimum: float | None = None,
    maximum: float | None = None,
) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise AssortedSourcesError(f"{context} is not numeric: {value!r}") from exc
    if not math.isfinite(number):
        raise AssortedSourcesError(f"{context} is not finite: {value!r}")
    if minimum is not None and number < minimum:
        raise AssortedSourcesError(f"{context} is below {minimum}: {number}")
    if maximum is not None and number > maximum:
        raise AssortedSourcesError(f"{context} is above {maximum}: {number}")
    return number


def _provenance(request: ArtifactRequest, checksum: str) -> dict[str, Any]:
    return {
        "source_id": request.source_id,
        "component_id": request.component_id,
        "source_version": request.source_version,
        "acquisition_url": request.url,
        "cached_file": str(request.cache_path),
        "cache_sha256": checksum,
    }


def normalize_nhtsa(
    request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    checksum: str | None = None,
) -> pd.DataFrame:
    """Extract NHTSA survival rates without aggregation or lifetime derivation."""
    digest = checksum or validate_cached_identity(request)
    member_expected = str(rules["workbook_member"])
    try:
        with ZipFile(request.cache_path) as archive:
            corrupt = archive.testzip()
            if corrupt is not None:
                raise AssortedSourcesError(
                    f"NHTSA ZIP contains a corrupt member: {corrupt}"
                )
            matches = [
                name
                for name in archive.namelist()
                if not name.endswith("/")
                and name.casefold() == member_expected.casefold()
            ]
            if len(matches) != 1:
                raise AssortedSourcesError(
                    f"NHTSA ZIP expected one {member_expected!r} member, got {matches}"
                )
            workbook_bytes = archive.read(matches[0])
    except BadZipFile as exc:
        raise AssortedSourcesError(f"NHTSA cache is not a valid ZIP: {request.cache_path}") from exc

    try:
        workbook = load_workbook(
            io.BytesIO(workbook_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise AssortedSourcesError(
            f"NHTSA member {member_expected} is not a readable XLSX workbook"
        ) from exc
    worksheet_name = str(rules["worksheet"])
    if worksheet_name not in workbook.sheetnames:
        workbook.close()
        raise AssortedSourcesError(
            f"NHTSA workbook missing worksheet {worksheet_name!r}; "
            f"available worksheets are {workbook.sheetnames}"
        )
    sheet = workbook[worksheet_name]
    for coordinate, expected in rules["title_headers"].items():
        actual = sheet[str(coordinate)].value
        if actual != expected:
            workbook.close()
            raise AssortedSourcesError(
                f"NHTSA header {coordinate} changed: expected {expected!r}, got {actual!r}"
            )

    source_range = str(rules["source_range"])
    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    data_start = int(rules["data_start_row"])
    spacer = int(rules["spacer_row"])
    if not (min_row <= spacer < data_start <= max_row):
        workbook.close()
        raise AssortedSourcesError("NHTSA configured header/data rows leave the source range")
    if any(
        sheet.cell(row=spacer, column=column).value is not None
        for column in range(min_col, max_col + 1)
    ):
        workbook.close()
        raise AssortedSourcesError(f"NHTSA expected blank spacer row {spacer}")

    class_header_row = int(rules["class_header_row"])
    class_columns = {
        str(column): str(label) for column, label in rules["class_columns"].items()
    }
    for column, expected in class_columns.items():
        actual = sheet[f"{column}{class_header_row}"].value
        if actual != expected:
            workbook.close()
            raise AssortedSourcesError(
                f"NHTSA vehicle-class label {column}{class_header_row} changed: "
                f"expected {expected!r}, got {actual!r}"
            )

    age_column = column_index_from_string(str(rules["age_column"]))
    expected_ages = list(
        range(
            int(rules["expected_ages"]["start"]),
            int(rules["expected_ages"]["end"]) + 1,
        )
    )
    actual_ages = [
        sheet.cell(row=row, column=age_column).value
        for row in range(data_start, max_row + 1)
    ]
    if actual_ages != expected_ages:
        workbook.close()
        raise AssortedSourcesError(
            f"NHTSA age coverage changed: expected {expected_ages}, got {actual_ages}"
        )

    bounds = rules["value_bounds"]
    records: list[dict[str, Any]] = []
    common = _provenance(request, digest)
    for source_row, age in zip(
        range(data_start, max_row + 1), expected_ages, strict=True
    ):
        for column, source_class in class_columns.items():
            coordinate = f"{column}{source_row}"
            value = _numeric(
                sheet[coordinate].value,
                context=f"NHTSA survival rate {coordinate}",
                minimum=float(bounds["minimum"]),
                maximum=float(bounds["maximum"]),
            )
            records.append(
                {
                    **common,
                    "source_vehicle_class_label": source_class,
                    "vehicle_age": age,
                    "survival_rate": value,
                    "unit": str(rules["unit"]),
                    "archive_member": matches[0],
                    "source_workbook": member_expected,
                    "source_worksheet": worksheet_name,
                    "source_range": source_range,
                    "source_cell": coordinate,
                }
            )
    workbook.close()
    frame = pd.DataFrame(records)
    key = ["source_vehicle_class_label", "vehicle_age"]
    if frame[key].isna().any().any() or frame.duplicated(key).any():
        raise AssortedSourcesError("NHTSA survival output contains null or duplicate keys")
    return frame.sort_values(key).reset_index(drop=True)


def normalize_nems(
    request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    checksum: str | None = None,
) -> pd.DataFrame:
    """Extract annual NEMS truck scrappage rates without converting to survival."""
    digest = checksum or validate_cached_identity(request)
    try:
        workbook = load_workbook(request.cache_path, read_only=True, data_only=True)
    except Exception as exc:
        raise AssortedSourcesError(
            f"NEMS cache is not a readable XLSX workbook: {request.cache_path}"
        ) from exc
    worksheet_name = str(rules["worksheet"])
    if worksheet_name not in workbook.sheetnames:
        workbook.close()
        raise AssortedSourcesError(
            f"NEMS workbook missing worksheet {worksheet_name!r}; "
            f"available worksheets are {workbook.sheetnames}"
        )
    sheet = workbook[worksheet_name]
    source_range = str(rules["source_range"])
    min_col, min_row, max_col, max_row = range_boundaries(source_range)
    actual_age_header = str(sheet.cell(row=min_row, column=min_col).value).strip()
    if actual_age_header != str(rules["age_header"]):
        workbook.close()
        raise AssortedSourcesError(
            f"NEMS age header changed: expected {rules['age_header']!r}, "
            f"got {actual_age_header!r}"
        )
    class_columns = {
        str(column): str(label) for column, label in rules["class_columns"].items()
    }
    for column, expected in class_columns.items():
        actual = sheet[f"{column}{min_row}"].value
        if actual != expected:
            workbook.close()
            raise AssortedSourcesError(
                f"NEMS class label {column}{min_row} changed: "
                f"expected {expected!r}, got {actual!r}"
            )
    expected_ages = list(
        range(
            int(rules["expected_ages"]["start"]),
            int(rules["expected_ages"]["end"]) + 1,
        )
    )
    actual_ages = [
        sheet.cell(row=row, column=min_col).value
        for row in range(min_row + 1, max_row + 1)
    ]
    if actual_ages != expected_ages:
        workbook.close()
        raise AssortedSourcesError(
            f"NEMS age coverage changed: expected {expected_ages}, got {actual_ages}"
        )
    bounds = rules["value_bounds"]
    common = _provenance(request, digest)
    records: list[dict[str, Any]] = []
    for source_row, age in zip(
        range(min_row + 1, max_row + 1), expected_ages, strict=True
    ):
        for column, source_class in class_columns.items():
            coordinate = f"{column}{source_row}"
            value = _numeric(
                sheet[coordinate].value,
                context=f"NEMS scrappage rate {coordinate}",
                minimum=float(bounds["minimum"]),
                maximum=float(bounds["maximum"]),
            )
            records.append(
                {
                    **common,
                    "source_vehicle_class_label": source_class,
                    "vehicle_age": age,
                    "annual_scrappage_rate": value,
                    "unit": str(rules["unit"]),
                    "source_workbook": str(rules["workbook"]),
                    "source_worksheet": worksheet_name,
                    "source_range": source_range,
                    "source_cell": coordinate,
                }
            )
    workbook.close()
    frame = pd.DataFrame(records)
    key = ["source_vehicle_class_label", "vehicle_age"]
    if frame[key].isna().any().any() or frame.duplicated(key).any():
        raise AssortedSourcesError("NEMS scrappage output contains null or duplicate keys")
    return frame.sort_values(key).reset_index(drop=True)


def normalize_gcam(
    request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    checksum: str | None = None,
) -> tuple[pd.DataFrame, SelectionStats]:
    """Select and reshape only the configured Canadian motorcycle source rows."""
    digest = checksum or validate_cached_identity(request)
    try:
        frame = pd.read_csv(
            request.cache_path, comment=str(rules["comment_prefix"])
        )
    except Exception as exc:
        raise AssortedSourcesError(f"Cannot read GCAM CSV: {request.cache_path}") from exc
    required = [str(column) for column in rules["required_columns"]]
    missing = sorted(set(required) - set(frame.columns))
    if missing:
        raise AssortedSourcesError(
            f"GCAM CSV missing required columns {missing}; available columns are "
            f"{list(frame.columns)}"
        )
    configured_years = [int(year) for year in rules["expected_years"]]
    actual_year_columns = [column for column in frame.columns if re.fullmatch(r"\d{4}", column)]
    if actual_year_columns != [str(year) for year in configured_years]:
        raise AssortedSourcesError(
            "GCAM year columns changed: "
            f"expected {configured_years}, got {actual_year_columns}"
        )

    selected = frame.copy()
    filters = {str(key): str(value) for key, value in rules["filters"].items()}
    for column, value in filters.items():
        selected = selected[selected[column].astype(str).eq(value)]
    selected = selected[selected["variable"].isin(rules["variables"])].copy()
    if selected.empty:
        raise AssortedSourcesError(
            f"GCAM exact filter selected no records: {filters}"
        )
    expected_variables = {str(value) for value in rules["variables"]}
    actual_variables = set(selected["variable"].astype(str))
    if actual_variables != expected_variables:
        raise AssortedSourcesError(
            f"GCAM variable set changed: expected {sorted(expected_variables)}, "
            f"got {sorted(actual_variables)}"
        )
    expected_technologies = {str(value) for value in rules["expected_technologies"]}
    actual_technologies = set(selected["UCD_technology"].astype(str))
    if actual_technologies != expected_technologies:
        raise AssortedSourcesError(
            f"GCAM technology set changed: expected {sorted(expected_technologies)}, "
            f"got {sorted(actual_technologies)}"
        )
    expected_units = {
        str(variable): str(unit) for variable, unit in rules["variable_units"].items()
    }
    for variable, unit in expected_units.items():
        actual_units = set(
            selected.loc[selected["variable"].eq(variable), "unit"].astype(str)
        )
        if actual_units != {unit}:
            raise AssortedSourcesError(
                f"GCAM unit for {variable!r} changed: expected {unit!r}, "
                f"got {sorted(actual_units)}"
            )

    source_rows = selected.reset_index(names="source_index")
    source_rows["source_row"] = source_rows["source_index"] + 2
    long = source_rows.melt(
        id_vars=[
            "source_row",
            "UCD_region",
            "UCD_sector",
            "mode",
            "size.class",
            "UCD_technology",
            "UCD_fuel",
            "variable",
            "unit",
        ],
        value_vars=[str(year) for year in configured_years],
        var_name="source_year",
        value_name="source_value",
    )
    long["source_year"] = pd.to_numeric(long["source_year"], errors="coerce")
    long["source_value"] = pd.to_numeric(long["source_value"], errors="coerce")
    if long[["source_year", "source_value"]].isna().any().any():
        raise AssortedSourcesError("GCAM selected rows contain null/non-numeric years or values")
    if (~long["source_value"].map(math.isfinite)).any():
        raise AssortedSourcesError("GCAM selected values contain non-finite numbers")
    long["source_year"] = long["source_year"].astype(int)
    duplicate_key = [str(value) for value in rules["duplicate_key"]]
    if long.duplicated(duplicate_key).any():
        raise AssortedSourcesError(
            f"GCAM selected records contain duplicate keys {duplicate_key}"
        )
    common = _provenance(request, digest)
    for column, value in common.items():
        long[column] = value
    long["source_filter"] = json.dumps(
        filters, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    )
    long = long.rename(
        columns={
            "UCD_region": "source_region",
            "UCD_sector": "source_sector",
            "mode": "source_mode",
            "size.class": "source_size_class",
            "UCD_technology": "source_technology",
            "UCD_fuel": "source_fuel",
            "variable": "source_variable",
            "unit": "source_unit",
        }
    )
    ordered = [
        *common,
        "source_region",
        "source_sector",
        "source_mode",
        "source_size_class",
        "source_technology",
        "source_fuel",
        "source_variable",
        "source_year",
        "source_value",
        "source_unit",
        "source_row",
        "source_filter",
    ]
    long = long[ordered].sort_values(
        ["source_variable", "source_technology", "source_year"]
    ).reset_index(drop=True)
    return long, SelectionStats(
        input_records=len(frame),
        selected_records=len(source_rows),
        excluded_records=len(frame) - len(source_rows),
    )


class _PageAssetParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_title = False
        self.title_parts: list[str] = []
        self.links: list[dict[str, str]] = []

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, str | None]]
    ) -> None:
        if tag.casefold() == "title":
            self.in_title = True
        if tag.casefold() == "link":
            self.links.append(
                {key.casefold(): value or "" for key, value in attrs}
            )

    def handle_endtag(self, tag: str) -> None:
        if tag.casefold() == "title":
            self.in_title = False

    def handle_data(self, data: str) -> None:
        if self.in_title:
            self.title_parts.append(data)


def discover_regen_payload_request(
    bundle: ConfigBundle,
    page_request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    checksum: str | None = None,
) -> tuple[ArtifactRequest, str]:
    """Resolve the content-hashed VuePress module linked by the cached page."""
    checksum or validate_cached_identity(page_request)
    try:
        html = page_request.cache_path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise AssortedSourcesError("REGEN page is not UTF-8 HTML") from exc
    parser = _PageAssetParser()
    parser.feed(html)
    actual_title = "".join(parser.title_parts).strip()
    if actual_title != str(rules["page_title"]):
        raise AssortedSourcesError(
            f"REGEN page title changed: expected {rules['page_title']!r}, "
            f"got {actual_title!r}"
        )
    selector = rules["payload_asset"]
    expected_rel = str(selector["rel"])
    expected_href = str(selector["href"])
    matches = []
    for link in parser.links:
        rel_values = link.get("rel", "").split()
        href = link.get("href", "")
        if expected_rel in rel_values and href == expected_href:
            matches.append(href)
    if len(matches) != 1:
        raise AssortedSourcesError(
            f"REGEN page expected one {expected_rel} {expected_href!r} asset, got {matches}"
        )
    source = _source(bundle, page_request.source_id)
    adapter = source.adapter
    payload_url = urljoin(str(adapter["access"]["asset_base_url"]), matches[0])
    request = ArtifactRequest(
        source_id=page_request.source_id,
        component_id=str(rules["component_id"]),
        source_version=page_request.source_version,
        url=payload_url,
        cache_path=resolve_input_path(
            bundle, "cache", str(adapter["payload_cache_path"])
        ),
        file_type="javascript",
        expected_sha256=str(adapter["expected_payload_sha256"]),
        expected_bytes=int(adapter["expected_payload_bytes"]),
    )
    return request, matches[0]


def _extract_js_assignment(payload: str, variable: str) -> str:
    pattern = re.compile(rf"(?:^|[,;]){re.escape(variable)}=")
    match = pattern.search(payload)
    if match is None:
        raise AssortedSourcesError(
            f"REGEN payload missing JavaScript assignment {variable!r}"
        )
    start = match.end()
    opening = payload[start : start + 1]
    closing = {"[": "]", "{": "}"}.get(opening)
    if closing is None:
        raise AssortedSourcesError(
            f"REGEN assignment {variable!r} is not a structured literal"
        )
    depth = 0
    quote: str | None = None
    escaped = False
    for index in range(start, len(payload)):
        character = payload[index]
        if quote is not None:
            if escaped:
                escaped = False
            elif character == "\\":
                escaped = True
            elif character == quote:
                quote = None
            continue
        if character in {'"', "'"}:
            quote = character
        elif character == opening:
            depth += 1
        elif character == closing:
            depth -= 1
            if depth == 0:
                return payload[start : index + 1]
    raise AssortedSourcesError(
        f"REGEN assignment {variable!r} has an unterminated literal"
    )


def _parse_js_literal(payload: str, variable: str) -> Any:
    literal = _extract_js_assignment(payload, variable)
    quoted = re.sub(
        r"([{,])([A-Za-z_$][A-Za-z0-9_$]*)\s*:",
        r'\1"\2":',
        literal,
    )
    try:
        return json.loads(quoted)
    except json.JSONDecodeError as exc:
        raise AssortedSourcesError(
            f"REGEN assignment {variable!r} is no longer a supported structured literal"
        ) from exc


def _measure_source(payload: str, field_variable: str) -> dict[str, Any]:
    literal = _extract_js_assignment(payload, field_variable)
    try:
        field_source = _parse_js_literal(payload, field_variable)
    except AssortedSourcesError:
        match = re.search(r"measures:([A-Za-z_$][A-Za-z0-9_$]*)", literal)
        if match is None:
            raise
        field_source = _parse_js_literal(payload, match.group(1))
    if "measures" in field_source:
        field_source = field_source["measures"]
    if not isinstance(field_source, dict):
        raise AssortedSourcesError(
            f"REGEN field source {field_variable!r} is not a measure mapping"
        )
    return field_source


def _measure_metadata(payload: str, field_variable: str, measure_field: str) -> dict[str, Any]:
    field_source = _measure_source(payload, field_variable)
    metadata = field_source.get(measure_field)
    if not isinstance(metadata, dict):
        raise AssortedSourcesError(
            f"REGEN field metadata missing measure {measure_field!r}"
        )
    return metadata


def normalize_regen(
    page_request: ArtifactRequest,
    payload_request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    payload_identity: str,
    page_checksum: str | None = None,
    payload_checksum: str | None = None,
) -> tuple[pd.DataFrame, list[dict[str, Any]], SelectionStats]:
    """Extract only the two configured Intercity Bus chart series."""
    page_digest = page_checksum or validate_cached_identity(page_request)
    payload_digest = payload_checksum or validate_cached_identity(payload_request)
    try:
        payload = payload_request.cache_path.read_text(encoding="utf-8")
    except UnicodeDecodeError as exc:
        raise AssortedSourcesError("REGEN payload is not UTF-8 JavaScript") from exc

    common = _provenance(payload_request, payload_digest)
    records: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    total_values = 0
    for chart_number, chart_rules in rules["charts"].items():
        chart_id = str(chart_rules["element_id"])
        chart_title = str(chart_rules["title"])
        call_pattern = re.compile(
            rf'id:"{re.escape(chart_id)}",caption:"{re.escape(chart_title)}",'
            rf'chartType:"[^"]+",dataSource:(?P<data>[A-Za-z_$][A-Za-z0-9_$]*),'
            rf'fieldSource:(?P<fields>[A-Za-z_$][A-Za-z0-9_$]*),'
            rf'keySource:(?P<keys>[A-Za-z_$][A-Za-z0-9_$]*),'
            rf'dimensionField:"{re.escape(str(rules["dimension_field"]))}",'
            rf'categoryField:"{re.escape(str(rules["category_field"]))}",'
            r'valueField:"(?P<measure>[A-Za-z_$][A-Za-z0-9_$]*)"}'
        )
        call = call_pattern.search(payload)
        if call is None:
            raise AssortedSourcesError(
                f"REGEN chart identity or payload structure changed: {chart_title}"
            )
        data = _parse_js_literal(payload, call.group("data"))
        keys = _parse_js_literal(payload, call.group("keys"))
        measure_field = str(chart_rules["measure_key"])
        metadata = _measure_metadata(payload, call.group("fields"), measure_field)
        if metadata.get("title") != str(rules["measure_title"]):
            raise AssortedSourcesError(
                f"REGEN {chart_title} measure changed: expected "
                f"{rules['measure_title']!r}, got {metadata.get('title')!r}"
            )
        unit = str(metadata.get("dataUnit", ""))
        if unit != str(chart_rules["expected_unit"]):
            raise AssortedSourcesError(
                f"REGEN {chart_title} unit changed: expected "
                f"{chart_rules['expected_unit']!r}, got {unit!r}"
            )
        expected_technologies = {
            str(key): str(label)
            for key, label in chart_rules["expected_technologies"].items()
        }
        actual_technologies = {
            str(item.get("Key")): str(item.get("Name"))
            for item in keys
            if item.get("Field") == str(rules["dimension_field"])
        }
        if actual_technologies != expected_technologies:
            raise AssortedSourcesError(
                f"REGEN {chart_title} technology metadata changed: expected "
                f"{expected_technologies}, got {actual_technologies}"
            )
        excluded_empty_series = {
            str(key): dict(specification)
            for key, specification in chart_rules.get(
                "excluded_empty_series", {}
            ).items()
        }
        unknown_exclusions = set(excluded_empty_series) - set(expected_technologies)
        if unknown_exclusions:
            raise AssortedSourcesError(
                f"REGEN {chart_title} config excludes unknown technologies: "
                f"{sorted(unknown_exclusions)}"
            )
        for technology_key, exclusion in excluded_empty_series.items():
            configured_label = str(exclusion.get("source_label", ""))
            expected_label = expected_technologies[technology_key]
            if configured_label != expected_label:
                raise AssortedSourcesError(
                    f"REGEN {chart_title} empty-series exclusion label changed: "
                    f"expected {expected_label!r}, got {configured_label!r}"
                )
            if not str(exclusion.get("reason", "")).strip():
                raise AssortedSourcesError(
                    f"REGEN {chart_title} empty-series exclusion for "
                    f"{technology_key!r} requires a reason"
                )
        expected_years = [int(year) for year in chart_rules["expected_years"]]
        actual_years = sorted({int(item[rules["category_field"]]) for item in data})
        if actual_years != expected_years:
            raise AssortedSourcesError(
                f"REGEN {chart_title} years changed: expected {expected_years}, "
                f"got {actual_years}"
            )
        lookup: dict[tuple[str, int], object] = {}
        for item in data:
            key = (
                str(item[rules["dimension_field"]]),
                int(item[rules["category_field"]]),
            )
            if key in lookup:
                raise AssortedSourcesError(
                    f"REGEN {chart_title} contains duplicate technology-year {key}"
                )
            lookup[key] = item.get(measure_field)
        total_values += len(data) * len(
            _measure_source(payload, call.group("fields"))
        )
        for technology_key, technology_label in expected_technologies.items():
            series_values: list[object] = []
            for year in expected_years:
                key = (technology_key, year)
                if key not in lookup:
                    raise AssortedSourcesError(
                        f"REGEN {chart_title} missing series record {key}"
                    )
                series_values.append(lookup[key])
            if technology_key in excluded_empty_series:
                if any(value is not None for value in series_values):
                    raise AssortedSourcesError(
                        f"REGEN {chart_title} configured empty-series exclusion "
                        f"{technology_label!r} contains a numeric value"
                    )
                exclusion = excluded_empty_series[technology_key]
                warnings.append(
                    {
                        "severity": "warning",
                        "code": "regen_configured_empty_series_excluded",
                        "source_id": payload_request.source_id,
                        "component_id": payload_request.component_id,
                        "chart_number": int(chart_number),
                        "chart_title": chart_title,
                        "measure": str(rules["measure_title"]),
                        "technology_key": technology_key,
                        "technology_label": technology_label,
                        "excluded_records": len(expected_years),
                        "reason": str(exclusion["reason"]),
                    }
                )
                continue
            for year, raw_value in zip(
                expected_years, series_values, strict=True
            ):
                value = _numeric(
                    raw_value,
                    context=(
                        f"REGEN {chart_title} {technology_label} {year} "
                        f"{rules['measure_title']}"
                    ),
                )
                records.append(
                    {
                        **common,
                        "page_acquisition_url": page_request.url,
                        "page_cached_file": str(page_request.cache_path),
                        "page_cache_sha256": page_digest,
                        "source_payload_identity": payload_identity,
                        "source_payload_element": (
                            f"{call.group('data')}.{measure_field}"
                        ),
                        "chart_number": int(chart_number),
                        "chart_id": chart_id,
                        "chart_title": chart_title,
                        "measure": str(rules["measure_title"]),
                        "metric": str(chart_rules["metric"]),
                        "source_technology_key": technology_key,
                        "source_technology_label": technology_label,
                        "source_year": year,
                        "source_value": value,
                        "native_unit": unit,
                        "currency": chart_rules.get("currency"),
                        "currency_year": chart_rules.get("currency_year"),
                    }
                )
        if str(chart_rules["metric"]) == "purchase_cost" and chart_rules.get(
            "currency_year"
        ) is None:
            warnings.append(
                {
                    "severity": "warning",
                    "code": "regen_currency_year_unresolved",
                    "source_id": payload_request.source_id,
                    "component_id": payload_request.component_id,
                    "chart_number": int(chart_number),
                    "message": (
                        "The REGEN page and structured payload label purchase costs "
                        "only with '$'; no currency year is explicit."
                    ),
                }
            )
    frame = pd.DataFrame(records)
    key = ["chart_number", "source_technology_key", "source_year"]
    if frame[key].isna().any().any() or frame.duplicated(key).any():
        raise AssortedSourcesError("REGEN output contains null or duplicate keys")
    frame = frame.sort_values(key).reset_index(drop=True)
    return frame, warnings, SelectionStats(
        input_records=total_values,
        selected_records=len(frame),
        excluded_records=max(total_values - len(frame), 0),
    )


def _normalize_whitespace(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def extract_pdf_text(
    request: ArtifactRequest,
    *,
    document_title: str,
    checksum: str | None = None,
) -> tuple[str, str]:
    """Validate a PDF and return its normal layout text layer."""
    digest = checksum or validate_cached_identity(request)
    try:
        reader = PdfReader(request.cache_path)
        if reader.is_encrypted and reader.decrypt("") == 0:
            raise AssortedSourcesError(f"FAA PDF is encrypted: {request.cache_path}")
        pages = [
            page.extract_text(extraction_mode="layout") or "" for page in reader.pages
        ]
    except (PdfReadError, OSError, ValueError) as exc:
        raise AssortedSourcesError(
            f"FAA PDF is not readable: {request.cache_path}"
        ) from exc
    text = "\n".join(pages)
    if len(_normalize_whitespace(text)) < 1000:
        raise AssortedSourcesError(
            f"FAA PDF text layer is empty or unusable: {request.cache_path}"
        )
    if _normalize_whitespace(document_title) not in _normalize_whitespace(text):
        raise AssortedSourcesError(
            f"FAA PDF missing document title {document_title!r}: {request.cache_path}"
        )
    return text, digest


def _parse_faa_value(raw: str, parse_method: str, *, context: str) -> tuple[float, str]:
    if parse_method == "percent":
        if not raw.endswith("%"):
            raise AssortedSourcesError(f"{context} lost its percent symbol: {raw!r}")
        return _numeric(raw[:-1].replace(",", ""), context=context), "strip_percent_keep_percent_unit"
    if parse_method == "currency":
        if not raw.startswith("$"):
            raise AssortedSourcesError(f"{context} lost its currency symbol: {raw!r}")
        return _numeric(raw[1:].replace(",", ""), context=context), "strip_currency_symbol"
    if parse_method == "numeric":
        if raw.startswith("$") or raw.endswith("%"):
            raise AssortedSourcesError(f"{context} has an unexpected symbol: {raw!r}")
        return _numeric(raw.replace(",", ""), context=context), "strip_grouping_separator"
    raise AssortedSourcesError(f"Unsupported FAA parse method {parse_method!r}")


def parse_faa_table_text(
    text: str,
    *,
    table_id: str,
    rules: dict[str, Any],
    request: ArtifactRequest,
    checksum: str,
) -> pd.DataFrame:
    """Parse one semantic FAA table and retain only configured All Aircraft metrics."""
    normalized = _normalize_whitespace(text)
    title = _normalize_whitespace(str(rules["title"]))
    start = normalized.find(title)
    if start < 0:
        raise AssortedSourcesError(f"FAA PDF missing expected table title {title!r}")
    source_marker = normalized.find(" Sources:", start)
    if source_marker < 0:
        source_marker = normalized.find(" Source:", start)
    if source_marker < 0:
        raise AssortedSourcesError(f"FAA Table {table_id} has no source boundary")
    segment = normalized[start:source_marker]
    for label in rules["required_header_labels"]:
        if _normalize_whitespace(str(label)) not in segment:
            raise AssortedSourcesError(
                f"FAA Table {table_id} missing metric/header label {label!r}"
            )
    category = str(rules["aircraft_category"])
    category_index = segment.rfind(category)
    if category_index < 0:
        raise AssortedSourcesError(
            f"FAA Table {table_id} missing aircraft category {category!r}"
        )
    row_text = segment[category_index + len(category) :].strip()
    tokens = re.findall(r"\$[\d,]+(?:\.\d+)?|[\d,]+(?:\.\d+)?%?", row_text)
    ordered_columns = [str(value) for value in rules["ordered_columns"]]
    if len(tokens) != len(ordered_columns):
        raise AssortedSourcesError(
            f"FAA Table {table_id} All Aircraft row is malformed: expected "
            f"{len(ordered_columns)} values, got {len(tokens)} ({tokens})"
        )
    raw_by_column = dict(zip(ordered_columns, tokens, strict=True))
    block_hours_raw = raw_by_column.get("block_hours")
    block_hours = (
        int(_numeric(block_hours_raw.replace(",", ""), context=f"FAA {table_id} block hours"))
        if block_hours_raw is not None
        else None
    )
    common = _provenance(request, checksum)
    records: list[dict[str, Any]] = []
    for source_column, selector in rules["selected_metrics"].items():
        source_column = str(source_column)
        if source_column not in raw_by_column:
            raise AssortedSourcesError(
                f"FAA Table {table_id} missing selected column {source_column!r}"
            )
        raw = raw_by_column[source_column]
        value, parse_method = _parse_faa_value(
            raw,
            str(selector["parse"]),
            context=f"FAA Table {table_id} {source_column}",
        )
        is_currency = str(selector["parse"]) == "currency"
        records.append(
            {
                **common,
                "source_component": request.component_id,
                "operating_group": str(rules["operating_group"]),
                "source_table": f"Table {table_id}",
                "source_table_title": str(rules["title"]),
                "aircraft_category": category,
                "metric": str(selector["metric"]),
                "value": value,
                "unit": str(selector["unit"]),
                "data_period": str(rules["data_period"]),
                "currency": str(rules["currency"]) if is_currency else pd.NA,
                "currency_year": rules.get("currency_year") if is_currency else pd.NA,
                "raw_value": raw,
                "parse_method": parse_method,
                "block_hours": block_hours,
            }
        )
    return pd.DataFrame(records)


def normalize_faa(
    section_3_request: ArtifactRequest,
    section_4_request: ArtifactRequest,
    rules: dict[str, Any],
    *,
    section_3_checksum: str | None = None,
    section_4_checksum: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]], SelectionStats]:
    """Extract requested FAA metrics and reconcile passenger/cargo categories."""
    section_3_text, digest_3 = extract_pdf_text(
        section_3_request,
        document_title=str(
            rules["documents"][section_3_request.component_id]["document_title"]
        ),
        checksum=section_3_checksum,
    )
    section_4_text, digest_4 = extract_pdf_text(
        section_4_request,
        document_title=str(
            rules["documents"][section_4_request.component_id]["document_title"]
        ),
        checksum=section_4_checksum,
    )
    tables: dict[str, pd.DataFrame] = {}
    for table_id in ("3-6", "3-9"):
        table_rules = {**rules, **rules["tables"][table_id]}
        tables[table_id] = parse_faa_table_text(
            section_3_text,
            table_id=table_id,
            rules=table_rules,
            request=section_3_request,
            checksum=digest_3,
        )
    for table_id in ("4-7", "4-8"):
        table_rules = {**rules, **rules["tables"][table_id]}
        tables[table_id] = parse_faa_table_text(
            section_4_text,
            table_id=table_id,
            rules=table_rules,
            request=section_4_request,
            checksum=digest_4,
        )
    for capacity_table, cost_table in (("3-6", "4-7"), ("3-9", "4-8")):
        capacity_categories = set(tables[capacity_table]["aircraft_category"])
        cost_categories = set(tables[cost_table]["aircraft_category"])
        if capacity_categories != cost_categories:
            raise AssortedSourcesError(
                f"FAA aircraft categories mismatch between Tables {capacity_table} "
                f"and {cost_table}: {capacity_categories} != {cost_categories}"
            )
    capacity = pd.concat([tables["3-6"], tables["3-9"]], ignore_index=True)
    maintenance = pd.concat([tables["4-7"], tables["4-8"]], ignore_index=True)
    capacity = capacity.sort_values(
        ["operating_group", "source_table", "metric"]
    ).reset_index(drop=True)
    maintenance = maintenance.sort_values(
        ["operating_group", "source_table", "metric"]
    ).reset_index(drop=True)
    key = ["operating_group", "source_table", "aircraft_category", "metric"]
    if capacity.duplicated(key).any() or maintenance.duplicated(key).any():
        raise AssortedSourcesError("FAA output contains duplicate metric keys")
    warnings = [
        {
            "severity": "warning",
            "code": "faa_currency_year_unresolved",
            "source_id": section_4_request.source_id,
            "component_id": section_4_request.component_id,
            "tables": ["Table 4-7", "Table 4-8"],
            "message": (
                "The FAA report identifies USD-denominated costs and YE June 2023 "
                "data, but does not explicitly establish a dollar year."
            ),
        }
    ]
    selected = len(capacity) + len(maintenance)
    all_aircraft_cells = sum(
        len(rules["tables"][table_id]["ordered_columns"])
        for table_id in ("3-6", "3-9", "4-7", "4-8")
    )
    return capacity, maintenance, warnings, SelectionStats(
        input_records=all_aircraft_cells,
        selected_records=selected,
        excluded_records=all_aircraft_cells - selected,
    )


def _manifest_row(
    request: ArtifactRequest,
    *,
    checksum: str,
    cache_status: str,
    stats: SelectionStats,
    selection: str,
    output_files: list[str],
    warning_count: int,
    status: str = "ok",
) -> dict[str, Any]:
    return {
        "source_id": request.source_id,
        "component_id": request.component_id,
        "source_version": request.source_version,
        "acquisition_url": request.url,
        "cached_file": str(request.cache_path),
        "sha256": checksum,
        "bytes": request.cache_path.stat().st_size,
        "cache_status": cache_status,
        "input_records": stats.input_records,
        "selected_records": stats.selected_records,
        "excluded_records": stats.excluded_records,
        "selection": selection,
        "output_files": "|".join(output_files),
        "warning_count": warning_count,
        "status": status,
    }


def write_outputs(
    *,
    outputs: dict[str, pd.DataFrame],
    manifest_rows: list[dict[str, Any]],
    warnings: list[dict[str, Any]],
    output_dir: Path,
    rules: dict[str, Any],
) -> None:
    """Write deterministic source-native CSV, manifest, and JSON-lines warnings."""
    output_dir.mkdir(parents=True, exist_ok=True)
    for filename in sorted(outputs):
        outputs[filename].to_csv(
            output_dir / filename, index=False, lineterminator="\n"
        )
    manifest = pd.DataFrame(manifest_rows, columns=MANIFEST_COLUMNS)
    manifest.to_csv(
        output_dir / str(rules["manifest_file"]), index=False, lineterminator="\n"
    )
    warning_lines = [
        json.dumps(warning, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        for warning in warnings
    ]
    (output_dir / str(rules["warnings_file"])).write_text(
        "\n".join(warning_lines) + ("\n" if warning_lines else ""),
        encoding="utf-8",
        newline="\n",
    )


def fetch_and_normalize(
    scenario_path: str | Path,
    *,
    download: bool = True,
    session: requests.Session | None = None,
) -> Path:
    """Fetch/cache every assorted source and write source-native interim artifacts."""
    bundle = load_config_bundle(scenario_path)
    rules = module_rules(bundle)
    requests_by_name = build_requests(bundle, rules)
    cache_status: dict[str, str] = {}
    checksums: dict[str, str] = {}
    for name, request in requests_by_name.items():
        cache_status[name] = ensure_cached(
            request, download=download, session=session
        )
        checksums[name] = validate_cached_identity(request)

    regen_rules = rules["epri_us_regen"]
    payload_request, payload_identity = discover_regen_payload_request(
        bundle,
        requests_by_name["regen_page"],
        regen_rules,
        checksum=checksums["regen_page"],
    )
    cache_status["regen_payload"] = ensure_cached(
        payload_request, download=download, session=session
    )
    checksums["regen_payload"] = validate_cached_identity(payload_request)

    nhtsa = normalize_nhtsa(
        requests_by_name["nhtsa"],
        rules["nhtsa_cafe"],
        checksum=checksums["nhtsa"],
    )
    nems = normalize_nems(
        requests_by_name["nems"],
        rules["eia_nems"],
        checksum=checksums["nems"],
    )
    gcam, gcam_stats = normalize_gcam(
        requests_by_name["gcam"],
        rules["jgcri_gcam"],
        checksum=checksums["gcam"],
    )
    regen, regen_warnings, regen_stats = normalize_regen(
        requests_by_name["regen_page"],
        payload_request,
        regen_rules,
        payload_identity=payload_identity,
        page_checksum=checksums["regen_page"],
        payload_checksum=checksums["regen_payload"],
    )
    capacity, maintenance, faa_warnings, faa_stats = normalize_faa(
        requests_by_name["faa_section_3"],
        requests_by_name["faa_section_4"],
        rules["faa"],
        section_3_checksum=checksums["faa_section_3"],
        section_4_checksum=checksums["faa_section_4"],
    )
    warnings = [*regen_warnings, *faa_warnings]
    outputs = {
        str(rules["nhtsa_cafe"]["output_file"]): nhtsa,
        str(rules["eia_nems"]["output_file"]): nems,
        str(rules["jgcri_gcam"]["output_file"]): gcam,
        str(regen_rules["output_file"]): regen,
        str(rules["faa"]["capacity_output_file"]): capacity,
        str(rules["faa"]["maintenance_output_file"]): maintenance,
    }

    manifest_rows = [
        _manifest_row(
            requests_by_name["nhtsa"],
            checksum=checksums["nhtsa"],
            cache_status=cache_status["nhtsa"],
            stats=SelectionStats(len(nhtsa), len(nhtsa), 0),
            selection="Vehicle Age Data!A3:E45; all configured classes and ages",
            output_files=[str(rules["nhtsa_cafe"]["output_file"])],
            warning_count=0,
        ),
        _manifest_row(
            requests_by_name["nems"],
            checksum=checksums["nems"],
            cache_status=cache_status["nems"],
            stats=SelectionStats(len(nems), len(nems), 0),
            selection="trnhdv!A86:D120; annual scrappage rates only",
            output_files=[str(rules["eia_nems"]["output_file"])],
            warning_count=0,
        ),
        _manifest_row(
            requests_by_name["gcam"],
            checksum=checksums["gcam"],
            cache_status=cache_status["gcam"],
            stats=gcam_stats,
            selection=json.dumps(
                rules["jgcri_gcam"]["filters"],
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ),
            output_files=[str(rules["jgcri_gcam"]["output_file"])],
            warning_count=0,
        ),
        _manifest_row(
            requests_by_name["regen_page"],
            checksum=checksums["regen_page"],
            cache_status=cache_status["regen_page"],
            stats=SelectionStats(1, 1, 0),
            selection=f"modulepreload {payload_identity}",
            output_files=[],
            warning_count=0,
        ),
        _manifest_row(
            payload_request,
            checksum=checksums["regen_payload"],
            cache_status=cache_status["regen_payload"],
            stats=regen_stats,
            selection="Charts 3 and 4; measure=Intercity Bus",
            output_files=[str(regen_rules["output_file"])],
            warning_count=len(regen_warnings),
            status="ok_with_warnings" if regen_warnings else "ok",
        ),
        _manifest_row(
            requests_by_name["faa_section_3"],
            checksum=checksums["faa_section_3"],
            cache_status=cache_status["faa_section_3"],
            stats=SelectionStats(
                input_records=sum(
                    len(rules["faa"]["tables"][table_id]["ordered_columns"])
                    for table_id in ("3-6", "3-9")
                ),
                selected_records=len(capacity),
                excluded_records=sum(
                    len(rules["faa"]["tables"][table_id]["ordered_columns"])
                    for table_id in ("3-6", "3-9")
                )
                - len(capacity),
            ),
            selection="Tables 3-6 and 3-9; All Aircraft requested metrics",
            output_files=[str(rules["faa"]["capacity_output_file"])],
            warning_count=0,
        ),
        _manifest_row(
            requests_by_name["faa_section_4"],
            checksum=checksums["faa_section_4"],
            cache_status=cache_status["faa_section_4"],
            stats=SelectionStats(
                input_records=faa_stats.input_records
                - sum(
                    len(rules["faa"]["tables"][table_id]["ordered_columns"])
                    for table_id in ("3-6", "3-9")
                ),
                selected_records=len(maintenance),
                excluded_records=faa_stats.excluded_records
                - (
                    sum(
                        len(rules["faa"]["tables"][table_id]["ordered_columns"])
                        for table_id in ("3-6", "3-9")
                    )
                    - len(capacity)
                ),
            ),
            selection="Tables 4-7 and 4-8; All Aircraft maintenance only",
            output_files=[str(rules["faa"]["maintenance_output_file"])],
            warning_count=len(faa_warnings),
            status="ok_with_warnings",
        ),
    ]
    output_dir = resolve_input_path(
        bundle, "interim", str(rules["interim_subdir"])
    )
    write_outputs(
        outputs=outputs,
        manifest_rows=manifest_rows,
        warnings=warnings,
        output_dir=output_dir,
        rules=rules,
    )
    return output_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--scenario",
        default="config/scenarios/legacy_reproduction.yaml",
        help="Scenario YAML used to resolve configured repository paths.",
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Require and reuse every configured cache artifact without network access.",
    )
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(message)s")
    args = parse_args()
    try:
        output_dir = fetch_and_normalize(
            args.scenario, download=not args.no_download
        )
    except (
        AssortedSourcesError,
        FileNotFoundError,
        requests.RequestException,
    ) as exc:
        raise SystemExit(f"Assorted-source adapter failed: {exc}") from exc
    logging.info("Wrote assorted-source interim outputs to %s", output_dir)


if __name__ == "__main__":
    main()
