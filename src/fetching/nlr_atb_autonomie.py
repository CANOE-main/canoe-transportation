"""Fetch, cache, and normalize 2024 Transportation ATB and 2022 ANL inputs."""

from __future__ import annotations

import argparse
import fnmatch
import hashlib
import io
import logging
import math
import os
import re
from pathlib import Path
from typing import Any, Literal
from urllib.parse import urlparse
from zipfile import BadZipFile, ZipFile

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, range_boundaries
from pydantic import BaseModel, ConfigDict, model_validator

from utils import (
    ConfigBundle,
    load_config_bundle,
    load_conversion_factors,
    load_harmonization_rules,
    resolve_input_path,
)
from validation.config_models import SourceSpec


ATB_SOURCE_ID = "nlr_atb_transportation_2024"
ANL_SOURCE_ID = "anl_autonomie_bean_2022"


class NlrAtbAutonomieError(ValueError):
    """Raised when an ATB/ANL source contract is unavailable or changed."""


class ArchiveComponentRequest(BaseModel):
    """Resolved contract for one required member of the ATB archive."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    component_id: str
    member_patterns: tuple[str, ...]
    file_format: Literal["csv", "xlsx"]
    normalizer: str
    required_columns: tuple[str, ...] = ()
    adapter: dict[str, Any]

    @model_validator(mode="after")
    def validate_request(self) -> "ArchiveComponentRequest":
        if not self.component_id or not self.member_patterns or not self.normalizer:
            raise ValueError("ATB component identity, member patterns, and normalizer are required")
        if any(Path(pattern).is_absolute() for pattern in self.member_patterns):
            raise ValueError("ATB member patterns must be archive-relative")
        return self


class AtbArchiveRequest(BaseModel):
    """Resolved source/cache/component contract for the official ATB ZIP."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    url: str
    cache_path: Path
    expected_trajectories: tuple[str, ...]
    components: tuple[ArchiveComponentRequest, ...]

    @model_validator(mode="after")
    def validate_request(self) -> "AtbArchiveRequest":
        parsed = urlparse(self.url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError(f"ATB URL is invalid: {self.url}")
        if not self.cache_path.is_absolute() or self.cache_path.suffix.casefold() != ".zip":
            raise ValueError(f"ATB cache must be an absolute .zip path: {self.cache_path}")
        if not self.expected_trajectories or not self.components:
            raise ValueError("ATB trajectories and archive components are required")
        return self


class ManualWorkbookRequest(BaseModel):
    """Resolved contract for the manually registered ANL workbook."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    folder_url: str
    workbook_path: Path
    sheet_name: str
    cell_range: str
    expected_columns: tuple[str, ...]
    table_layout: dict[str, Any]
    output_file: str
    required: bool = False

    @model_validator(mode="after")
    def validate_request(self) -> "ManualWorkbookRequest":
        parsed = urlparse(self.folder_url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError(f"ANL Box URL is invalid: {self.folder_url}")
        if not self.workbook_path.is_absolute() or self.workbook_path.suffix.casefold() not in {
            ".xlsx",
            ".xlsm",
        }:
            raise ValueError(
                f"ANL workbook must be an absolute .xlsx/.xlsm path: {self.workbook_path}"
            )
        if (
            not self.sheet_name
            or not self.cell_range
            or not self.expected_columns
            or not self.table_layout
        ):
            raise ValueError(
                "ANL workbook sheet, range, columns, and table layout are required"
            )
        min_col, min_row, max_col, max_row = range_boundaries(self.cell_range)
        if min_col > max_col or min_row > max_row:
            raise ValueError(f"ANL workbook range is invalid: {self.cell_range}")
        if max_col - min_col + 1 != len(self.expected_columns):
            raise ValueError(
                "ANL expected columns must match the configured workbook range width"
            )
        actual_columns = tuple(
            chr(ord("A") + index) for index in range(min_col - 1, max_col)
        )
        if self.expected_columns != actual_columns:
            raise ValueError(
                "ANL expected columns must identify the configured workbook range"
            )
        return self


def module_rules(bundle: ConfigBundle) -> dict[str, Any]:
    """Load ATB/ANL normalization and output rules."""
    return load_harmonization_rules(bundle, "nlr_atb_autonomie")


def _source(bundle: ConfigBundle, source_id: str) -> SourceSpec:
    source = bundle.sources.sources.get(source_id)
    if not isinstance(source, SourceSpec):
        raise NlrAtbAutonomieError(f"sources.yaml missing {source_id}")
    return source


def configured_trajectory(bundle: ConfigBundle) -> str:
    """Resolve the scenario-selected ATB trajectory used only as an output marker."""
    source = _source(bundle, ATB_SOURCE_ID)
    expected = tuple(str(value) for value in source.adapter["expected_trajectories"])
    selection = bundle.scenario.sources.selections.get(ATB_SOURCE_ID)
    trajectory = selection.trajectory if selection is not None else None
    if trajectory is None:
        raise NlrAtbAutonomieError(
            f"Scenario must select a trajectory for {ATB_SOURCE_ID}"
        )
    if trajectory not in expected:
        raise NlrAtbAutonomieError(
            f"Unsupported ATB trajectory {trajectory!r}; expected one of {list(expected)}"
        )
    return trajectory


def build_atb_request(bundle: ConfigBundle) -> AtbArchiveRequest:
    """Build the configured official ATB archive request."""
    source = _source(bundle, ATB_SOURCE_ID)
    access = source.adapter.get("access", {})
    component_requests: list[ArchiveComponentRequest] = []
    for component_id, component in source.components.items():
        adapter = dict(component.adapter)
        component_requests.append(
            ArchiveComponentRequest(
                component_id=str(component_id),
                member_patterns=tuple(str(value) for value in adapter.pop("member_patterns")),
                file_format=str(adapter.pop("format")),
                normalizer=str(adapter.pop("normalizer")),
                required_columns=tuple(
                    str(value) for value in adapter.pop("required_columns", [])
                ),
                adapter=adapter,
            )
        )
    return AtbArchiveRequest(
        url=str(access.get("url", "")),
        cache_path=resolve_input_path(
            bundle, "cache", str(source.adapter.get("cache_path", ""))
        ),
        expected_trajectories=tuple(
            str(value) for value in source.adapter.get("expected_trajectories", [])
        ),
        components=tuple(component_requests),
    )


def build_manual_request(
    bundle: ConfigBundle, rules: dict[str, Any] | None = None
) -> ManualWorkbookRequest:
    """Build the registered manual ANL workbook request."""
    source = _source(bundle, ANL_SOURCE_ID)
    component = source.component("mhdv_maintenance_coefficients")
    adapter = source.adapter
    access = adapter.get("access", {})
    root = resolve_input_path(bundle, "external", str(adapter["external_subdir"]))
    selected_rules = rules or module_rules(bundle)
    workbook_rules = selected_rules["components"]["anl_bean"]
    return ManualWorkbookRequest(
        folder_url=str(access.get("folder_url", "")),
        workbook_path=root / str(adapter["expected_workbook"]),
        sheet_name=str(workbook_rules["workbook_sheet"]),
        cell_range=str(workbook_rules["workbook_range"]),
        expected_columns=tuple(
            str(value) for value in workbook_rules["expected_columns"]
        ),
        table_layout=dict(workbook_rules["table_layout"]),
        output_file=str(workbook_rules["output_file"]),
        required=bool(source.required or component.required),
    )


def fetch_archive_to_cache(
    request: AtbArchiveRequest,
    *,
    session: requests.Session | None = None,
    timeout: int = 120,
) -> str:
    """Download the official ATB ZIP atomically, or reuse an existing cache."""
    if request.cache_path.is_file():
        return "cached"
    request.cache_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = request.cache_path.with_suffix(request.cache_path.suffix + ".part")
    client = session or requests.Session()
    try:
        response = client.get(request.url, stream=True, timeout=timeout)
        response.raise_for_status()
        with temporary.open("wb") as handle:
            for chunk in response.iter_content(chunk_size=1024 * 1024):
                if chunk:
                    handle.write(chunk)
        if not temporary.is_file() or temporary.stat().st_size == 0:
            raise NlrAtbAutonomieError(f"ATB download was empty: {request.url}")
        try:
            with ZipFile(temporary) as archive:
                if archive.testzip() is not None:
                    raise NlrAtbAutonomieError("ATB download contains a corrupt ZIP member")
        except BadZipFile as exc:
            raise NlrAtbAutonomieError(
                f"ATB download is not a valid ZIP: {request.url}"
            ) from exc
        os.replace(temporary, request.cache_path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return "downloaded"


def discover_zip_members(
    archive_path: Path,
    components: tuple[ArchiveComponentRequest, ...],
) -> dict[str, str]:
    """Find one case-insensitive archive member for each configured component."""
    if not archive_path.is_file():
        raise FileNotFoundError(archive_path)
    try:
        with ZipFile(archive_path) as archive:
            names = [name for name in archive.namelist() if not name.endswith("/")]
    except BadZipFile as exc:
        raise NlrAtbAutonomieError(f"ATB cache is not a valid ZIP: {archive_path}") from exc

    discovered: dict[str, str] = {}
    for component in components:
        matches = {
            name
            for name in names
            if any(
                _matches_archive_member(name, pattern)
                for pattern in component.member_patterns
            )
        }
        if not matches:
            raise NlrAtbAutonomieError(
                f"ATB archive missing required component {component.component_id}; "
                f"expected one of {list(component.member_patterns)}"
            )
        if len(matches) > 1:
            raise NlrAtbAutonomieError(
                f"ATB archive component {component.component_id} is ambiguous: "
                f"{sorted(matches)}"
            )
        discovered[component.component_id] = matches.pop()
    return discovered


def _matches_archive_member(name: str, pattern: str) -> bool:
    """Match an exact member or the same member below one release directory."""
    normalized_name = name.replace("\\", "/").casefold()
    normalized_pattern = pattern.replace("\\", "/").casefold()
    if normalized_pattern.startswith("*/"):
        suffix = normalized_pattern[2:]
        if len(normalized_name.split("/")) != len(suffix.split("/")) + 1:
            return False
    return fnmatch.fnmatchcase(normalized_name, normalized_pattern)


def _require_columns(
    frame: pd.DataFrame, required: tuple[str, ...] | list[str], context: str
) -> None:
    missing = sorted(set(required) - set(frame.columns))
    if missing:
        raise NlrAtbAutonomieError(
            f"{context} missing required columns {missing}; available columns are "
            f"{list(frame.columns)}"
        )
    if frame.empty:
        raise NlrAtbAutonomieError(f"{context} contains no rows")


def normalize_vehicles(
    frame: pd.DataFrame,
    *,
    request: AtbArchiveRequest,
    component: ArchiveComponentRequest,
    source_member: str,
    default_trajectory: str,
    rules: dict[str, Any],
) -> pd.DataFrame:
    """Retain configured cost/performance metrics and every ATB trajectory."""
    _require_columns(frame, component.required_columns, "ATB vehicles.csv")
    actual_trajectories = set(frame["scenario"].dropna().astype(str))
    expected_trajectories = set(request.expected_trajectories)
    if actual_trajectories != expected_trajectories:
        raise NlrAtbAutonomieError(
            "ATB vehicle trajectories changed: "
            f"expected {sorted(expected_trajectories)}, got {sorted(actual_trajectories)}"
        )
    vehicle_rules = rules["components"]["vehicles"]
    metric_keys = {str(key): str(value) for key, value in vehicle_rules["metric_keys"].items()}
    metric_units = {
        str(key): str(value) for key, value in vehicle_rules["metric_units"].items()
    }
    selected = frame[frame["metric"].isin(metric_keys)].copy()
    missing_metrics = sorted(set(metric_keys) - set(selected["metric"].astype(str)))
    if missing_metrics:
        raise NlrAtbAutonomieError(
            f"ATB vehicles.csv missing configured metrics {missing_metrics}"
        )
    phev_powertrain = str(vehicle_rules["phev_powertrain"])
    phev_reconciliation_metrics = {
        str(metric) for metric in vehicle_rules["phev_reconciliation_metrics"].values()
    }
    is_phev_fuel_economy = selected["vehicle_powertrain"].eq(
        phev_powertrain
    ) & selected["metric"].isin(phev_reconciliation_metrics)
    selected = selected.loc[~is_phev_fuel_economy].copy()
    selected["year"] = pd.to_numeric(selected["year"], errors="coerce")
    selected["value"] = pd.to_numeric(selected["value"], errors="coerce")
    if selected[["year", "value"]].isna().any().any():
        raise NlrAtbAutonomieError("ATB vehicles.csv contains invalid year/value rows")
    selected["year"] = selected["year"].astype(int)
    selected["trajectory"] = selected["scenario"].astype(str)
    selected["metric_key"] = selected["metric"].map(metric_keys)
    selected["unit"] = selected["metric"].map(metric_units)
    selected["is_default_trajectory"] = selected["trajectory"].eq(default_trajectory)
    selected["source_id"] = ATB_SOURCE_ID
    selected["source_member"] = source_member
    return selected.drop(columns=["scenario"]).sort_values(
        ["vehicle_class", "vehicle_powertrain", "vehicle_detail", "metric", "trajectory", "year"],
        na_position="last",
    ).reset_index(drop=True)


def _numeric_columns(
    frame: pd.DataFrame,
    columns: list[str],
    *,
    context: str,
) -> pd.DataFrame:
    normalized = frame.copy()
    for column in columns:
        normalized[column] = pd.to_numeric(normalized[column], errors="coerce")
    invalid = normalized[columns].isna() | ~normalized[columns].map(
        lambda value: math.isfinite(float(value))
    )
    if invalid.any().any():
        bad_columns = [column for column in columns if invalid[column].any()]
        raise NlrAtbAutonomieError(
            f"{context} contains non-numeric or non-finite values in {bad_columns}"
        )
    return normalized


def combine_mhdv_phev_cycles(
    frame: pd.DataFrame,
    *,
    rules: dict[str, Any],
) -> pd.DataFrame:
    """Combine active MHDV cycles using consumption-correct averaging."""
    cycle_rules = rules["cycle_aggregation"]["cycles"]
    contribution_columns = [
        str(cycle["contribution"]) for cycle in cycle_rules.values()
    ]
    numeric = _numeric_columns(
        frame,
        contribution_columns,
        context="ATB MHDV PHEV cycle inputs",
    )
    contributions = numeric[contribution_columns]
    if (contributions < 0).any().any():
        raise NlrAtbAutonomieError(
            "ATB MHDV PHEV cycle contributions must be non-negative"
        )
    tolerance = float(rules["tolerances"]["cycle_contribution_sum"])
    sums = contributions.sum(axis=1)
    if ((sums - 1.0).abs() > tolerance).any():
        raise NlrAtbAutonomieError(
            "ATB MHDV PHEV active cycle contributions must sum to 1 within "
            f"{tolerance}"
        )

    reciprocal_consumption = pd.Series(0.0, index=numeric.index)
    combined_electricity = pd.Series(0.0, index=numeric.index)
    for cycle in cycle_rules.values():
        weight = numeric[str(cycle["contribution"])]
        cs = pd.to_numeric(
            numeric[str(cycle["cs_fuel_economy"])], errors="coerce"
        )
        cd = pd.to_numeric(
            numeric[str(cycle["cd_electricity_consumption"])], errors="coerce"
        )
        active = weight > 0
        active_values = pd.concat([cs[active], cd[active]], axis=1)
        invalid_active = active_values.isna() | ~active_values.map(
            lambda value: math.isfinite(float(value))
        )
        if invalid_active.any().any():
            raise NlrAtbAutonomieError(
                "ATB MHDV PHEV active-cycle CS/CD values must be numeric and finite"
            )
        if (cs[active] <= 0).any():
            raise NlrAtbAutonomieError(
                "ATB MHDV PHEV active-cycle CS fuel economy must be positive"
            )
        if (cd[active] < 0).any():
            raise NlrAtbAutonomieError(
                "ATB MHDV PHEV active-cycle CD electricity consumption "
                "must be non-negative"
            )
        reciprocal_consumption.loc[active] += weight[active] / cs[active]
        combined_electricity.loc[active] += weight[active] * cd[active]
    if (reciprocal_consumption <= 0).any():
        raise NlrAtbAutonomieError(
            "ATB MHDV PHEV cycle aggregation produced non-positive fuel consumption"
        )
    return pd.DataFrame(
        {
            "combined_cs_fuel_economy_mi_per_gallon_equivalent": (
                1.0 / reciprocal_consumption
            ),
            "combined_cd_electricity_consumption_wh_per_mi": combined_electricity,
            "source_cycle_contribution_sum": sums,
        },
        index=numeric.index,
    )


def _validate_utility_factor_table(
    frame: pd.DataFrame,
    *,
    key_fields: list[str],
    utility_factor_column: str,
    context: str,
    range_column: str | None = None,
) -> pd.DataFrame:
    _require_columns(
        frame,
        [*key_fields, utility_factor_column],
        context,
    )
    numeric_columns = [utility_factor_column]
    if range_column is not None:
        numeric_columns.append(range_column)
    normalized = _numeric_columns(frame, numeric_columns, context=context)
    if normalized.duplicated(key_fields, keep=False).any():
        duplicates = normalized.loc[
            normalized.duplicated(key_fields, keep=False), key_fields
        ].drop_duplicates()
        raise NlrAtbAutonomieError(
            f"{context} has ambiguous duplicate keys: "
            f"{duplicates.to_dict(orient='records')[:5]}"
        )
    if (
        (normalized[utility_factor_column] < 0)
        | (normalized[utility_factor_column] > 1)
    ).any():
        raise NlrAtbAutonomieError(f"{context} utility factors must lie in [0, 1]")
    return normalized


def _joined_text(values: pd.Series) -> str:
    return " | ".join(dict.fromkeys(str(value) for value in values.dropna() if str(value)))


def match_phev_utility_factors(
    vehicles: pd.DataFrame,
    *,
    ldv_utility_factors: pd.DataFrame,
    mdhd_utility_factors: pd.DataFrame,
    rules: dict[str, Any],
) -> pd.DataFrame:
    """Match exact MHDV UFs and exact or bounded-interpolated LDV UFs."""
    matching = rules["utility_factor_matching"]
    uf_column = str(matching["utility_factor_column"])
    reference_column = str(matching["reference_column"])
    notes_column = str(matching["notes_column"])
    ldv_rules = matching["ldv"]
    mdhd_rules = matching["mdhd"]
    ldv_keys = [str(value) for value in ldv_rules["exact_key_fields"]]
    mdhd_keys = [str(value) for value in mdhd_rules["exact_key_fields"]]
    range_column = str(ldv_rules["interpolation_dimension"])

    ldv_table = _validate_utility_factor_table(
        ldv_utility_factors,
        key_fields=ldv_keys,
        utility_factor_column=uf_column,
        range_column=range_column,
        context="ATB LDV PHEV utility-factor table",
    )
    mdhd_table = _validate_utility_factor_table(
        mdhd_utility_factors,
        key_fields=mdhd_keys,
        utility_factor_column=uf_column,
        context="ATB MHDV PHEV utility-factor table",
    )

    result = vehicles.copy()
    match_rows: list[dict[str, Any]] = []
    ldv_weight = str(rules["light_duty_weight_category"])
    mdhd_weight = str(rules["medium_heavy_weight_category"])
    for _, vehicle in result.iterrows():
        weight_category = str(vehicle["vehicle_weight_category"])
        if weight_category == ldv_weight:
            target_range = float(vehicle["electric_range_mi"])
            candidates = ldv_table[
                ldv_table["vehicle_weight_category"].astype(str).eq(weight_category)
            ].sort_values(range_column)
            exact = candidates[candidates[range_column].eq(target_range)]
            if len(exact) == 1:
                matched = exact.iloc[0]
                match_rows.append(
                    {
                        "fleet_utility_factor": float(matched[uf_column]),
                        "utility_factor_source_family": str(
                            ldv_rules["source_family"]
                        ),
                        "utility_factor_match_method": "exact",
                        "utility_factor_matched_range_mi": target_range,
                        "utility_factor_lower_range_mi": target_range,
                        "utility_factor_upper_range_mi": target_range,
                        "utility_factor_reference": str(matched[reference_column]),
                        "utility_factor_notes": str(matched[notes_column]),
                    }
                )
                continue
            if len(exact) > 1:
                raise NlrAtbAutonomieError(
                    f"Ambiguous LDV PHEV utility factor for range {target_range}"
                )
            if str(ldv_rules["interpolation_method"]) != "linear":
                raise NlrAtbAutonomieError(
                    f"No exact LDV PHEV utility factor for range {target_range}"
                )
            lower = candidates[candidates[range_column] < target_range].tail(1)
            upper = candidates[candidates[range_column] > target_range].head(1)
            if lower.empty or upper.empty:
                raise NlrAtbAutonomieError(
                    "LDV PHEV utility-factor interpolation would extrapolate for "
                    f"{weight_category}, range {target_range}"
                )
            low = lower.iloc[0]
            high = upper.iloc[0]
            low_range = float(low[range_column])
            high_range = float(high[range_column])
            fraction = (target_range - low_range) / (high_range - low_range)
            utility_factor = float(low[uf_column]) + fraction * (
                float(high[uf_column]) - float(low[uf_column])
            )
            match_rows.append(
                {
                    "fleet_utility_factor": utility_factor,
                    "utility_factor_source_family": str(ldv_rules["source_family"]),
                    "utility_factor_match_method": "linear_interpolation",
                    "utility_factor_matched_range_mi": target_range,
                    "utility_factor_lower_range_mi": low_range,
                    "utility_factor_upper_range_mi": high_range,
                    "utility_factor_reference": _joined_text(
                        pd.Series([low[reference_column], high[reference_column]])
                    ),
                    "utility_factor_notes": _joined_text(
                        pd.Series([low[notes_column], high[notes_column]])
                    ),
                }
            )
        elif weight_category == mdhd_weight:
            candidates = mdhd_table
            for key in mdhd_keys:
                candidates = candidates[
                    candidates[key].astype(str).eq(str(vehicle[key]))
                ]
            if len(candidates) != 1:
                raise NlrAtbAutonomieError(
                    "MHDV PHEV utility-factor match must produce exactly one row for "
                    f"{[(key, vehicle[key]) for key in mdhd_keys]}; got {len(candidates)}"
                )
            matched = candidates.iloc[0]
            match_rows.append(
                {
                    "fleet_utility_factor": float(matched[uf_column]),
                    "utility_factor_source_family": str(mdhd_rules["source_family"]),
                    "utility_factor_match_method": "exact",
                    "utility_factor_matched_range_mi": pd.NA,
                    "utility_factor_lower_range_mi": pd.NA,
                    "utility_factor_upper_range_mi": pd.NA,
                    "utility_factor_reference": str(matched[reference_column]),
                    "utility_factor_notes": str(matched[notes_column]),
                }
            )
        else:
            raise NlrAtbAutonomieError(
                f"Unsupported PHEV vehicle weight category {weight_category!r}"
            )

    matches = pd.DataFrame(match_rows, index=result.index)
    result = pd.concat([result, matches], axis=1)
    tolerance = float(rules["tolerances"]["utility_factor_bounds"])
    if (
        (result["fleet_utility_factor"] < -tolerance)
        | (result["fleet_utility_factor"] > 1.0 + tolerance)
    ).any():
        raise NlrAtbAutonomieError("Matched PHEV utility factors fall outside [0, 1]")
    return result


def _conversion_value(
    conversions: dict[str, Any],
    rules: dict[str, Any],
    name: str,
) -> float:
    value: Any = conversions
    path = [str(part) for part in rules["conversion_keys"][name]]
    for part in path:
        if not isinstance(value, dict) or part not in value:
            raise NlrAtbAutonomieError(
                f"Missing configured conversion factor {'.'.join(path)}"
            )
        value = value[part]
    try:
        numeric = float(value)
    except (TypeError, ValueError) as exc:
        raise NlrAtbAutonomieError(
            f"Configured conversion factor {'.'.join(path)} is not numeric"
        ) from exc
    if not math.isfinite(numeric) or numeric <= 0:
        raise NlrAtbAutonomieError(
            f"Configured conversion factor {'.'.join(path)} must be positive and finite"
        )
    return numeric


def _reconcile_phev_total_fuel_economy(
    derived: pd.DataFrame,
    *,
    output_vehicles: pd.DataFrame,
    rules: dict[str, Any],
) -> pd.DataFrame:
    reconciliation = rules["reconciliation"]
    key_fields = [str(value) for value in reconciliation["output_key_fields"]]
    derived_key_fields = {
        str(output): str(derived)
        for output, derived in reconciliation["derived_to_output_key_fields"].items()
    }
    if set(derived_key_fields) != set(key_fields):
        raise NlrAtbAutonomieError(
            "PHEV reconciliation derived/output key mapping is incomplete"
        )
    metric_by_basis = {
        str(key): str(value)
        for key, value in reconciliation["metric_by_basis"].items()
    }
    evidence = output_vehicles[
        output_vehicles["vehicle_powertrain"].eq(str(rules["source_powertrain"]))
        & output_vehicles["metric"].isin(metric_by_basis.values())
    ].copy()
    evidence["year"] = pd.to_numeric(evidence["year"], errors="coerce")
    evidence["value"] = pd.to_numeric(evidence["value"], errors="coerce")
    if evidence[["year", "value"]].isna().any().any():
        raise NlrAtbAutonomieError(
            "ATB output PHEV reconciliation rows contain invalid year/value data"
        )
    evidence["year"] = evidence["year"].astype(int)
    evidence_key_fields = [*key_fields, "metric"]
    if evidence.duplicated(evidence_key_fields, keep=False).any():
        raise NlrAtbAutonomieError(
            "ATB output PHEV fuel economy has ambiguous reconciliation keys"
        )
    evidence_lookup = {
        tuple(row[field] for field in evidence_key_fields): row
        for _, row in evidence.iterrows()
    }
    aliases = {
        str(key): str(value)
        for key, value in reconciliation.get("scenario_aliases", {}).items()
    }
    rows: list[dict[str, Any]] = []
    for _, row in derived.iterrows():
        basis = str(row["fuel_equivalent_basis"])
        metric = metric_by_basis[basis]
        key_values = {
            field: row[derived_key_fields[field]] for field in key_fields
        }
        exact_key = tuple([*(key_values[field] for field in key_fields), metric])
        matched = evidence_lookup.get(exact_key)
        method = "exact"
        matched_scenario = str(row["trajectory"])
        if matched is None and str(row["trajectory"]) in aliases:
            matched_scenario = aliases[str(row["trajectory"])]
            key_values["scenario"] = matched_scenario
            alias_key = tuple([*(key_values[field] for field in key_fields), metric])
            matched = evidence_lookup.get(alias_key)
            method = "scenario_alias" if matched is not None else "missing"
        elif matched is None:
            method = "missing"
        if matched is None:
            rows.append(
                {
                    "reconciliation_output_scenario": pd.NA,
                    "reconciliation_match_method": method,
                    "reconciliation_output_metric": metric,
                    "reconciliation_output_fuel_economy_mi_per_gallon_equivalent": pd.NA,
                    "reconciliation_absolute_difference": pd.NA,
                    "reconciliation_relative_difference": pd.NA,
                    "reconciliation_within_tolerance": pd.NA,
                }
            )
            continue
        published = float(matched["value"])
        calculated = float(
            row["reconciliation_derived_total_fuel_economy_mi_per_gallon_equivalent"]
        )
        absolute_difference = abs(calculated - published)
        relative_difference = absolute_difference / abs(published) if published else math.inf
        within = math.isclose(
            calculated,
            published,
            rel_tol=float(reconciliation["relative_tolerance"]),
            abs_tol=float(
                reconciliation["absolute_tolerance_mi_per_gallon_equivalent"]
            ),
        )
        rows.append(
            {
                "reconciliation_output_scenario": matched_scenario,
                "reconciliation_match_method": method,
                "reconciliation_output_metric": metric,
                "reconciliation_output_fuel_economy_mi_per_gallon_equivalent": published,
                "reconciliation_absolute_difference": absolute_difference,
                "reconciliation_relative_difference": relative_difference,
                "reconciliation_within_tolerance": within,
            }
        )
    return pd.concat(
        [derived.reset_index(drop=True), pd.DataFrame(rows)], axis=1
    )


def derive_phev_efficiency(
    vehicle_inputs: pd.DataFrame,
    *,
    ldv_utility_factors: pd.DataFrame,
    mdhd_utility_factors: pd.DataFrame,
    output_vehicles: pd.DataFrame,
    rules: dict[str, Any],
    conversions: dict[str, Any],
    source_members: dict[str, str],
    default_trajectory: str,
) -> pd.DataFrame:
    """Derive one auditable utility-weighted efficiency row per ATB PHEV input."""
    powertrain = str(rules["source_powertrain"])
    selected = vehicle_inputs[
        vehicle_inputs["vehicle_powertrain"].astype(str).eq(powertrain)
    ].copy()
    if selected.empty:
        raise NlrAtbAutonomieError("ATB inputs_vehicles.csv contains no PHEV rows")
    key_fields = [str(value) for value in rules["source_key_fields"]]
    if selected.duplicated(key_fields, keep=False).any():
        raise NlrAtbAutonomieError(
            "ATB PHEV vehicle inputs have ambiguous source-dimension keys"
        )
    selected["source_vehicle_input_row"] = selected.index + 2
    selected = _numeric_columns(
        selected,
        ["year", "range(mi)"],
        context="ATB PHEV vehicle inputs",
    )
    selected["year"] = selected["year"].astype(int)

    source_fields = {
        str(source): str(target)
        for target, source in rules["source_fields"].items()
    }
    selected = selected.rename(columns=source_fields)
    selected["trajectory"] = selected["source_scenario"].astype(str)
    cycle_rename: dict[str, str] = {}
    for cycle_name, cycle in rules["cycle_aggregation"]["cycles"].items():
        prefix = str(cycle_name).casefold()
        cycle_rename[str(cycle["contribution"])] = (
            f"source_{prefix}_cycle_contribution"
        )
        cycle_rename[str(cycle["cs_fuel_economy"])] = (
            f"source_{prefix}_cs_fuel_economy_mi_per_dge"
        )
        cycle_rename[str(cycle["cd_electricity_consumption"])] = (
            f"source_{prefix}_cd_electricity_consumption_wh_per_mi"
        )

    ldv_weight = str(rules["light_duty_weight_category"])
    mdhd_weight = str(rules["medium_heavy_weight_category"])
    ldv = selected["vehicle_weight_category"].eq(ldv_weight)
    mdhd = selected["vehicle_weight_category"].eq(mdhd_weight)
    if (~(ldv | mdhd)).any():
        unexpected = sorted(
            selected.loc[~(ldv | mdhd), "vehicle_weight_category"].astype(str).unique()
        )
        raise NlrAtbAutonomieError(
            f"ATB PHEV inputs contain unsupported weight categories {unexpected}"
        )

    selected[
        "combined_cs_fuel_economy_mi_per_gallon_equivalent"
    ] = pd.NA
    selected["combined_cd_electricity_consumption_wh_per_mi"] = pd.NA
    selected["source_cycle_contribution_sum"] = pd.NA
    ldv_values = _numeric_columns(
        selected.loc[
            ldv,
            [
                "source_combined_cs_fuel_economy_mi_per_gge",
                "source_combined_cd_electricity_consumption_wh_per_mi",
            ],
        ],
        [
            "source_combined_cs_fuel_economy_mi_per_gge",
            "source_combined_cd_electricity_consumption_wh_per_mi",
        ],
        context="ATB LDV PHEV combined inputs",
    )
    if (ldv_values["source_combined_cs_fuel_economy_mi_per_gge"] <= 0).any():
        raise NlrAtbAutonomieError("ATB LDV PHEV combined CS fuel economy must be positive")
    if (
        ldv_values["source_combined_cd_electricity_consumption_wh_per_mi"] < 0
    ).any():
        raise NlrAtbAutonomieError(
            "ATB LDV PHEV combined CD electricity consumption must be non-negative"
        )
    selected.loc[
        ldv, "combined_cs_fuel_economy_mi_per_gallon_equivalent"
    ] = ldv_values["source_combined_cs_fuel_economy_mi_per_gge"]
    selected.loc[
        ldv, "combined_cd_electricity_consumption_wh_per_mi"
    ] = ldv_values["source_combined_cd_electricity_consumption_wh_per_mi"]

    raw_cycle_rules = rules["cycle_aggregation"]["cycles"]
    pre_rename_cycle_columns = [
        str(value)
        for cycle in raw_cycle_rules.values()
        for value in cycle.values()
    ]
    original_mdhd = vehicle_inputs.loc[selected.index[mdhd]]
    cycle_combined = combine_mhdv_phev_cycles(
        original_mdhd[pre_rename_cycle_columns],
        rules=rules,
    )
    for column in cycle_combined:
        selected.loc[mdhd, column] = cycle_combined[column]
    selected = selected.rename(columns=cycle_rename)

    basis_map = {
        str(key): str(value) for key, value in rules["fuel_equivalent_basis"].items()
    }
    selected["fuel_equivalent_basis"] = selected["secondary_fuel"].map(basis_map)
    if selected["fuel_equivalent_basis"].isna().any():
        fuels = sorted(
            selected.loc[
                selected["fuel_equivalent_basis"].isna(), "secondary_fuel"
            ].astype(str).unique()
        )
        raise NlrAtbAutonomieError(
            f"ATB PHEV secondary fuels have no equivalent-gallon basis: {fuels}"
        )

    selected = match_phev_utility_factors(
        selected,
        ldv_utility_factors=ldv_utility_factors,
        mdhd_utility_factors=mdhd_utility_factors,
        rules=rules,
    )
    selected[
        "combined_cs_fuel_economy_mi_per_gallon_equivalent"
    ] = pd.to_numeric(
        selected["combined_cs_fuel_economy_mi_per_gallon_equivalent"]
    )
    selected[
        "combined_cd_electricity_consumption_wh_per_mi"
    ] = pd.to_numeric(
        selected["combined_cd_electricity_consumption_wh_per_mi"]
    )
    utility_factor = selected["fleet_utility_factor"]
    selected[
        "utility_weighted_fuel_consumption_gallon_equivalent_per_mi"
    ] = (1.0 - utility_factor) / selected[
        "combined_cs_fuel_economy_mi_per_gallon_equivalent"
    ]
    selected[
        "utility_weighted_electricity_consumption_wh_per_mi"
    ] = utility_factor * selected[
        "combined_cd_electricity_consumption_wh_per_mi"
    ]

    mile_to_km = _conversion_value(conversions, rules, "mile_to_km")
    us_gallon_to_litre = _conversion_value(
        conversions, rules, "us_gallon_to_litre"
    )
    wh_to_kwh = _conversion_value(conversions, rules, "wh_to_kwh")
    fuel_conversion = _conversion_value(
        conversions,
        rules,
        "gal_equivalent_per_mile_to_litre_equivalent_per_100_km",
    )
    electricity_conversion = _conversion_value(
        conversions, rules, "wh_per_mile_to_kwh_per_100_km"
    )
    expected_fuel_conversion = us_gallon_to_litre * 100.0 / mile_to_km
    expected_electricity_conversion = 100.0 * wh_to_kwh / mile_to_km
    if not math.isclose(
        fuel_conversion, expected_fuel_conversion, rel_tol=1.0e-15, abs_tol=0.0
    ):
        raise NlrAtbAutonomieError(
            "Configured gal-equivalent/mi to litre-equivalent/100 km factor "
            "does not match its configured primitives"
        )
    if not math.isclose(
        electricity_conversion,
        expected_electricity_conversion,
        rel_tol=1.0e-15,
        abs_tol=0.0,
    ):
        raise NlrAtbAutonomieError(
            "Configured Wh/mi to kWh/100 km factor does not match its "
            "configured primitives"
        )
    wh_per_gge = _conversion_value(conversions, rules, "wh_per_gge")
    wh_per_dge = _conversion_value(conversions, rules, "wh_per_dge")
    selected["wh_per_fuel_equivalent_gallon"] = selected[
        "fuel_equivalent_basis"
    ].map({"gge": wh_per_gge, "dge": wh_per_dge})
    selected[
        "utility_weighted_fuel_consumption_litre_equivalent_per_100_km"
    ] = (
        selected["utility_weighted_fuel_consumption_gallon_equivalent_per_mi"]
        * fuel_conversion
    )
    selected[
        "utility_weighted_electricity_consumption_kwh_per_100_km"
    ] = (
        selected["utility_weighted_electricity_consumption_wh_per_mi"]
        * electricity_conversion
    )
    selected["utility_weighted_fuel_energy_wh_equivalent_per_mi"] = (
        selected["utility_weighted_fuel_consumption_gallon_equivalent_per_mi"]
        * selected["wh_per_fuel_equivalent_gallon"]
    )
    selected["total_utility_weighted_energy_wh_equivalent_per_mi"] = (
        selected["utility_weighted_fuel_energy_wh_equivalent_per_mi"]
        + selected["utility_weighted_electricity_consumption_wh_per_mi"]
    )
    if (
        selected["total_utility_weighted_energy_wh_equivalent_per_mi"] <= 0
    ).any():
        raise NlrAtbAutonomieError(
            "ATB PHEV total utility-weighted energy must be positive"
        )
    selected["electricity_input_share"] = (
        selected["utility_weighted_electricity_consumption_wh_per_mi"]
        / selected["total_utility_weighted_energy_wh_equivalent_per_mi"]
    )
    selected["liquid_fuel_input_share"] = 1.0 - selected["electricity_input_share"]
    share_tolerance = float(rules["tolerances"]["input_share_sum"])
    if (
        (selected["electricity_input_share"] < -share_tolerance)
        | (selected["electricity_input_share"] > 1.0 + share_tolerance)
        | (selected["liquid_fuel_input_share"] < -share_tolerance)
        | (selected["liquid_fuel_input_share"] > 1.0 + share_tolerance)
        | (
            (
                selected["electricity_input_share"]
                + selected["liquid_fuel_input_share"]
                - 1.0
            ).abs()
            > share_tolerance
        )
    ).any():
        raise NlrAtbAutonomieError(
            "ATB PHEV energy input shares must lie in [0,1] and sum to 1"
        )

    units = rules["units"]
    selected["combined_cs_fuel_economy_unit"] = selected[
        "fuel_equivalent_basis"
    ].map(units["combined_cs"])
    selected["utility_weighted_fuel_consumption_source_unit"] = selected[
        "fuel_equivalent_basis"
    ].map(units["utility_weighted_fuel_source"])
    selected["utility_weighted_fuel_consumption_canadian_unit"] = selected[
        "fuel_equivalent_basis"
    ].map(units["utility_weighted_fuel_canadian"])
    selected["utility_weighted_electricity_consumption_source_unit"] = str(
        units["utility_weighted_electricity_source"]
    )
    selected["utility_weighted_electricity_consumption_canadian_unit"] = str(
        units["utility_weighted_electricity_canadian"]
    )
    selected[
        "reconciliation_derived_total_fuel_economy_mi_per_gallon_equivalent"
    ] = 1.0 / (
        selected["utility_weighted_fuel_consumption_gallon_equivalent_per_mi"]
        + selected["utility_weighted_electricity_consumption_wh_per_mi"]
        / selected["wh_per_fuel_equivalent_gallon"]
    )
    selected["trajectory"] = selected["trajectory"].astype(str)
    aliases = {
        str(key): str(value)
        for key, value in rules["reconciliation"].get("scenario_aliases", {}).items()
    }
    selected["is_default_trajectory"] = selected["trajectory"].map(
        lambda value: aliases.get(value, value) == default_trajectory
    )
    selected["source_id"] = ATB_SOURCE_ID
    selected["source_vehicle_input_member"] = source_members["phev_vehicle_inputs"]
    selected["source_utility_factor_member"] = selected[
        "utility_factor_source_family"
    ].map(
        {
            str(rules["utility_factor_matching"]["ldv"]["source_family"]): (
                source_members["phev_utility_factor_ldv"]
            ),
            str(rules["utility_factor_matching"]["mdhd"]["source_family"]): (
                source_members["phev_utility_factor_mdhd"]
            ),
        }
    )
    selected["source_reconciliation_member"] = source_members["vehicles"]
    selected = _reconcile_phev_total_fuel_economy(
        selected,
        output_vehicles=output_vehicles,
        rules=rules,
    )
    return selected.sort_values(
        [
            "vehicle_weight_category",
            "vehicle_class",
            "vehicle_detail",
            "trajectory",
            "year",
        ]
    ).reset_index(drop=True)


def normalize_vmt_ldv(
    frame: pd.DataFrame,
    *,
    component: ArchiveComponentRequest,
    source_member: str,
    rules: dict[str, Any],
) -> pd.DataFrame:
    """Normalize the row-oriented LDV age/VMT schedule."""
    _require_columns(frame, component.required_columns, "ATB vmt_ldv.csv")
    value_column = str(rules["components"]["vmt"]["value_column"])
    normalized = frame.rename(columns={"vmt(mi)": value_column}).copy()
    normalized["year_index"] = pd.to_numeric(normalized["year_index"], errors="coerce")
    normalized[value_column] = pd.to_numeric(normalized[value_column], errors="coerce")
    if normalized[["year_index", value_column]].isna().any().any():
        raise NlrAtbAutonomieError("ATB vmt_ldv.csv contains invalid age/VMT rows")
    normalized["year_index"] = normalized["year_index"].astype(int)
    normalized["vmt_source_component"] = component.component_id
    normalized["unit"] = str(rules["components"]["vmt"]["unit"])
    normalized["source_id"] = ATB_SOURCE_ID
    normalized["source_member"] = source_member
    return normalized


def normalize_vmt_mdhd(
    frame: pd.DataFrame,
    *,
    component: ArchiveComponentRequest,
    source_member: str,
    rules: dict[str, Any],
) -> pd.DataFrame:
    """Unpivot numeric MD/HD age columns into the common age/VMT shape."""
    _require_columns(frame, component.required_columns, "ATB vmt_mdhd.csv")
    pattern = re.compile(str(component.adapter["age_column_pattern"]))
    age_columns = [str(column) for column in frame.columns if pattern.fullmatch(str(column))]
    if not age_columns:
        raise NlrAtbAutonomieError("ATB vmt_mdhd.csv has no configured age columns")
    ages = sorted(int(column) for column in age_columns)
    if ages != list(range(ages[0], ages[-1] + 1)):
        raise NlrAtbAutonomieError(
            f"ATB vmt_mdhd.csv age columns are not contiguous: {ages}"
        )
    id_columns = [column for column in frame.columns if str(column) not in age_columns]
    value_column = str(rules["components"]["vmt"]["value_column"])
    normalized = frame.melt(
        id_vars=id_columns,
        value_vars=age_columns,
        var_name="year_index",
        value_name=value_column,
    )
    normalized["year_index"] = pd.to_numeric(normalized["year_index"], errors="coerce")
    normalized[value_column] = pd.to_numeric(normalized[value_column], errors="coerce")
    if normalized[["year_index", value_column]].isna().any().any():
        raise NlrAtbAutonomieError("ATB vmt_mdhd.csv contains invalid age/VMT rows")
    normalized["year_index"] = normalized["year_index"].astype(int)
    normalized["vmt_source_component"] = component.component_id
    normalized["unit"] = str(rules["components"]["vmt"]["unit"])
    normalized["source_id"] = ATB_SOURCE_ID
    normalized["source_member"] = source_member
    return normalized


def read_ldv_maintenance_workbook(
    content: bytes,
    *,
    component: ArchiveComponentRequest,
    source_member: str,
) -> dict[str, pd.DataFrame]:
    """Read and validate every configured LDV maintenance worksheet."""
    sheets = component.adapter.get("sheets")
    if not isinstance(sheets, dict) or not sheets:
        raise NlrAtbAutonomieError("ATB LDV maintenance sheet contracts are missing")
    workbook = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
    missing_sheets = sorted(set(sheets) - set(workbook.sheet_names))
    if missing_sheets:
        raise NlrAtbAutonomieError(
            f"ATB maintenance_ldv.xlsx missing worksheets {missing_sheets}"
        )
    outputs: dict[str, pd.DataFrame] = {}
    for sheet_name, sheet_rules in sheets.items():
        frame = pd.read_excel(workbook, sheet_name=str(sheet_name))
        _require_columns(
            frame,
            [str(value) for value in sheet_rules["required_columns"]],
            f"ATB maintenance_ldv.xlsx[{sheet_name}]",
        )
        frame["workbook_sheet"] = str(sheet_name)
        frame["source_id"] = ATB_SOURCE_ID
        frame["source_member"] = source_member
        outputs[str(sheet_name)] = frame
    return outputs


def extract_bean_coefficients(request: ManualWorkbookRequest) -> pd.DataFrame:
    """Extract the configured ANL coefficient blocks to a source-labelled long table."""
    if not request.workbook_path.is_file():
        raise FileNotFoundError(
            f"Required manual ANL workbook missing: {request.workbook_path}. "
            f"Download the complete Box folder from {request.folder_url} into "
            f"{request.workbook_path.parent}."
        )
    try:
        workbook = load_workbook(
            request.workbook_path,
            read_only=True,
            data_only=True,
            keep_vba=request.workbook_path.suffix.casefold() == ".xlsm",
        )
    except Exception as exc:
        raise NlrAtbAutonomieError(
            f"Unable to read ANL workbook {request.workbook_path}: {exc}"
        ) from exc
    if request.sheet_name not in workbook.sheetnames:
        workbook.close()
        raise NlrAtbAutonomieError(
            f"ANL workbook missing worksheet {request.sheet_name!r}; "
            f"available worksheets are {workbook.sheetnames}"
        )
    min_col, min_row, max_col, max_row = range_boundaries(request.cell_range)
    worksheet = workbook[request.sheet_name]
    layout = request.table_layout
    header_row = int(layout["header_row"])
    label_column = str(layout["label_column"])
    powertrain_column = str(layout["powertrain_column"])
    vehicle_columns = {
        str(column): str(label)
        for column, label in dict(layout["vehicle_class_columns"]).items()
    }
    expected_powertrains = [str(value) for value in layout["expected_powertrains"]]
    if not min_row <= header_row <= max_row:
        workbook.close()
        raise NlrAtbAutonomieError("ANL header row falls outside the configured range")

    for column, expected_label in vehicle_columns.items():
        column_index = column_index_from_string(column)
        if not min_col <= column_index <= max_col:
            workbook.close()
            raise NlrAtbAutonomieError(
                f"ANL vehicle-class column {column} falls outside {request.cell_range}"
            )
        actual_label = worksheet[f"{column}{header_row}"].value
        if actual_label != expected_label:
            workbook.close()
            raise NlrAtbAutonomieError(
                f"ANL vehicle-class header changed at {column}{header_row}: "
                f"expected {expected_label!r}, got {actual_label!r}"
            )

    records: list[dict[str, Any]] = []
    for coefficient_key, block in dict(layout["blocks"]).items():
        first_row, last_row = (int(value) for value in block["rows"])
        if first_row < min_row or last_row > max_row or first_row > last_row:
            workbook.close()
            raise NlrAtbAutonomieError(
                f"ANL block {coefficient_key} falls outside {request.cell_range}"
            )
        source_label = str(block["source_label"])
        actual_label = worksheet[f"{label_column}{first_row}"].value
        if actual_label != source_label:
            workbook.close()
            raise NlrAtbAutonomieError(
                f"ANL coefficient label changed at {label_column}{first_row}: "
                f"expected {source_label!r}, got {actual_label!r}"
            )
        actual_powertrains = [
            worksheet[f"{powertrain_column}{row}"].value
            for row in range(first_row, last_row + 1)
        ]
        if actual_powertrains != expected_powertrains:
            workbook.close()
            raise NlrAtbAutonomieError(
                f"ANL powertrains changed for {coefficient_key}: "
                f"expected {expected_powertrains}, got {actual_powertrains}"
            )
        for row, powertrain in zip(
            range(first_row, last_row + 1),
            actual_powertrains,
            strict=True,
        ):
            for column, vehicle_class in vehicle_columns.items():
                value = worksheet[f"{column}{row}"].value
                try:
                    numeric_value = float(value)
                except (TypeError, ValueError) as exc:
                    workbook.close()
                    raise NlrAtbAutonomieError(
                        f"ANL coefficient is not numeric at {column}{row}: {value!r}"
                    ) from exc
                records.append(
                    {
                        "coefficient_key": str(coefficient_key),
                        "source_coefficient_label": source_label,
                        "vehicle_powertrain": str(powertrain),
                        "vehicle_class": vehicle_class,
                        "value": numeric_value,
                        "source_row": row,
                        "source_column": column,
                    }
                )
    workbook.close()
    frame = pd.DataFrame(records)
    expected_count = (
        len(layout["blocks"])
        * len(expected_powertrains)
        * len(vehicle_columns)
    )
    if len(frame) != expected_count:
        raise NlrAtbAutonomieError(
            f"ANL coefficient extraction expected {expected_count} rows, got {len(frame)}"
        )
    frame["source_id"] = ANL_SOURCE_ID
    frame["source_workbook"] = str(request.workbook_path)
    frame["source_sheet"] = request.sheet_name
    frame["source_range"] = request.cell_range
    return frame


def file_sha256(path: Path) -> str:
    """Return a stable SHA-256 checksum for one physical source artifact."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_outputs(
    *,
    outputs: dict[str, pd.DataFrame],
    manifest_rows: list[dict[str, Any]],
    warnings: list[str],
    output_dir: Path,
    rules: dict[str, Any],
) -> None:
    """Write normalized source tables, manifest, and warnings."""
    output_dir.mkdir(parents=True, exist_ok=True)
    for filename, frame in outputs.items():
        frame.to_csv(output_dir / filename, index=False)
    pd.DataFrame(manifest_rows).to_csv(output_dir / rules["manifest_file"], index=False)
    (output_dir / rules["warnings_file"]).write_text(
        "\n".join(warnings) + ("\n" if warnings else ""),
        encoding="utf-8",
    )


def fetch_and_normalize(
    scenario_path: str | Path,
    *,
    download: bool = True,
    session: requests.Session | None = None,
) -> Path:
    """Fetch/cache ATB and write source-normalized ATB/available ANL outputs."""
    bundle = load_config_bundle(scenario_path)
    rules = module_rules(bundle)
    conversions = load_conversion_factors(bundle)
    request = build_atb_request(bundle)
    default_trajectory = configured_trajectory(bundle)
    output_dir = resolve_input_path(bundle, "interim", str(rules["interim_subdir"]))

    if download:
        cache_status = fetch_archive_to_cache(request, session=session)
    elif request.cache_path.is_file():
        cache_status = "cached"
    else:
        raise FileNotFoundError(
            f"ATB offline cache missing: {request.cache_path}. Run without --no-download "
            "or place the official ZIP at the configured cache path."
        )

    members = discover_zip_members(request.cache_path, request.components)
    archive_checksum = file_sha256(request.cache_path)
    outputs: dict[str, pd.DataFrame] = {}
    vmt_frames: list[pd.DataFrame] = []
    phev_frames: dict[str, pd.DataFrame] = {}
    phev_source_members: dict[str, str] = {}
    output_vehicle_evidence: pd.DataFrame | None = None
    manifest_rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    with ZipFile(request.cache_path) as archive:
        for component in request.components:
            member = members[component.component_id]
            normalizer = component.normalizer
            input_rows = 0
            output_files: list[str] = []
            selected_rows = 0
            if component.file_format == "xlsx":
                sheet_frames = read_ldv_maintenance_workbook(
                    archive.read(member), component=component, source_member=member
                )
                configured_outputs = rules["components"]["maintenance_ldv"]["outputs"]
                for sheet_name, frame in sheet_frames.items():
                    filename = str(configured_outputs[sheet_name])
                    outputs[filename] = frame
                    output_files.append(filename)
                    input_rows += len(frame)
                    selected_rows += len(frame)
            else:
                frame = pd.read_csv(archive.open(member))
                input_rows = len(frame)
                if normalizer == "vehicles":
                    output_vehicle_evidence = frame.copy()
                    normalized = normalize_vehicles(
                        frame,
                        request=request,
                        component=component,
                        source_member=member,
                        default_trajectory=default_trajectory,
                        rules=rules,
                    )
                    filename = str(rules["components"]["vehicles"]["output_file"])
                    outputs[filename] = normalized
                    output_files.append(filename)
                elif normalizer in {
                    "phev_vehicle_inputs",
                    "phev_utility_factor_ldv",
                    "phev_utility_factor_mdhd",
                }:
                    _require_columns(
                        frame,
                        component.required_columns,
                        f"ATB {component.component_id}",
                    )
                    if normalizer == "phev_vehicle_inputs":
                        normalized = frame[
                            frame["vehicle_powertrain"]
                            .astype(str)
                            .eq(
                                str(
                                    rules["components"]["phev_efficiency"][
                                        "source_powertrain"
                                    ]
                                )
                            )
                        ].copy()
                    else:
                        normalized = frame.copy()
                    if normalized.empty:
                        raise NlrAtbAutonomieError(
                            f"ATB {component.component_id} selected no rows"
                        )
                    phev_frames[normalizer] = frame
                    phev_source_members[normalizer] = member
                    output_files.append(
                        str(rules["components"]["phev_efficiency"]["output_file"])
                    )
                elif normalizer == "vmt_ldv":
                    normalized = normalize_vmt_ldv(
                        frame,
                        component=component,
                        source_member=member,
                        rules=rules,
                    )
                    vmt_frames.append(normalized)
                    output_files.append(str(rules["components"]["vmt"]["output_file"]))
                elif normalizer == "vmt_mdhd":
                    normalized = normalize_vmt_mdhd(
                        frame,
                        component=component,
                        source_member=member,
                        rules=rules,
                    )
                    vmt_frames.append(normalized)
                    output_files.append(str(rules["components"]["vmt"]["output_file"]))
                else:
                    raise NlrAtbAutonomieError(
                        f"No ATB normalizer configured for {normalizer!r}"
                    )
                selected_rows = len(normalized)
            manifest_rows.append(
                {
                    "source_id": ATB_SOURCE_ID,
                    "component_id": component.component_id,
                    "source_url": request.url,
                    "cached_file": str(request.cache_path),
                    "sha256": archive_checksum,
                    "cache_status": cache_status,
                    "source_member": member,
                    "input_rows": input_rows,
                    "selected_rows": selected_rows,
                    "output_files": "|".join(output_files),
                    "default_trajectory": default_trajectory,
                    "status": "ok",
                }
            )

    required_phev_frames = {
        "phev_vehicle_inputs",
        "phev_utility_factor_ldv",
        "phev_utility_factor_mdhd",
    }
    missing_phev_frames = sorted(required_phev_frames - set(phev_frames))
    if missing_phev_frames or output_vehicle_evidence is None:
        raise NlrAtbAutonomieError(
            "ATB archive did not provide the complete PHEV derivation contract: "
            f"missing {missing_phev_frames or ['vehicles']}"
        )
    phev_rules = rules["components"]["phev_efficiency"]
    phev_source_members["vehicles"] = members["vehicles"]
    phev = derive_phev_efficiency(
        phev_frames["phev_vehicle_inputs"],
        ldv_utility_factors=phev_frames["phev_utility_factor_ldv"],
        mdhd_utility_factors=phev_frames["phev_utility_factor_mdhd"],
        output_vehicles=output_vehicle_evidence,
        rules=phev_rules,
        conversions=conversions,
        source_members=phev_source_members,
        default_trajectory=default_trajectory,
    )
    phev_filename = str(phev_rules["output_file"])
    outputs[phev_filename] = phev
    missing_reconciliation = int(
        phev["reconciliation_output_fuel_economy_mi_per_gallon_equivalent"]
        .isna()
        .sum()
    )
    outside_tolerance = int(
        phev["reconciliation_within_tolerance"].eq(False).sum()  # noqa: E712
    )
    if missing_reconciliation or outside_tolerance:
        comparable = phev["reconciliation_relative_difference"].dropna()
        max_relative = float(comparable.max()) if not comparable.empty else math.nan
        warnings.append(
            "PHEV output fuel-economy reconciliation is report-only: "
            f"{missing_reconciliation} missing and {outside_tolerance} outside "
            f"configured tolerance across {len(phev)} derived rows; maximum relative "
            f"difference={max_relative:.12g}."
        )
    manifest_rows.append(
        {
            "source_id": ATB_SOURCE_ID,
            "component_id": "phev_efficiency_derivation",
            "source_url": request.url,
            "cached_file": str(request.cache_path),
            "sha256": archive_checksum,
            "cache_status": cache_status,
            "source_member": "|".join(
                phev_source_members[key]
                for key in (
                    "phev_vehicle_inputs",
                    "phev_utility_factor_ldv",
                    "phev_utility_factor_mdhd",
                    "vehicles",
                )
            ),
            "input_rows": len(phev_frames["phev_vehicle_inputs"]),
            "selected_rows": len(phev),
            "output_files": phev_filename,
            "default_trajectory": default_trajectory,
            "status": "ok_with_reconciliation_differences"
            if missing_reconciliation or outside_tolerance
            else "ok",
        }
    )

    if not vmt_frames:
        raise NlrAtbAutonomieError("ATB archive produced no VMT schedules")
    vmt_filename = str(rules["components"]["vmt"]["output_file"])
    outputs[vmt_filename] = pd.concat(vmt_frames, ignore_index=True, sort=False).sort_values(
        ["vehicle_weight_category", "vehicle_class", "year_index"],
        na_position="last",
    ).reset_index(drop=True)

    manual = build_manual_request(bundle, rules)
    if manual.workbook_path.is_file():
        coefficients = extract_bean_coefficients(manual)
        outputs[manual.output_file] = coefficients
        _, manual_min_row, _, manual_max_row = range_boundaries(manual.cell_range)
        manifest_rows.append(
            {
                "source_id": ANL_SOURCE_ID,
                "component_id": "mhdv_maintenance_coefficients",
                "source_url": manual.folder_url,
                "cached_file": str(manual.workbook_path),
                "sha256": file_sha256(manual.workbook_path),
                "cache_status": "manual",
                "source_member": f"{manual.sheet_name}!{manual.cell_range}",
                "input_rows": manual_max_row - manual_min_row + 1,
                "selected_rows": len(coefficients),
                "output_files": manual.output_file,
                "default_trajectory": "",
                "status": "ok",
            }
        )
    else:
        message = (
            f"ANL manual input unavailable: {manual.workbook_path}. Download the complete "
            f"Box folder from {manual.folder_url} into {manual.workbook_path.parent}."
        )
        if manual.required:
            raise FileNotFoundError(message)
        warnings.append(message)
        manifest_rows.append(
            {
                "source_id": ANL_SOURCE_ID,
                "component_id": "mhdv_maintenance_coefficients",
                "source_url": manual.folder_url,
                "cached_file": str(manual.workbook_path),
                "sha256": "",
                "cache_status": "manual_missing",
                "source_member": f"{manual.sheet_name}!{manual.cell_range}",
                "input_rows": 0,
                "selected_rows": 0,
                "output_files": "",
                "default_trajectory": "",
                "status": "warning",
            }
        )

    write_outputs(
        outputs=outputs,
        manifest_rows=manifest_rows,
        warnings=warnings,
        output_dir=output_dir,
        rules=rules,
    )
    return output_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--scenario",
        default="config/scenarios/legacy_reproduction.yaml",
        help="Scenario YAML controlling configured paths and the default trajectory marker.",
    )
    parser.add_argument(
        "--no-download",
        action="store_true",
        help="Require and reuse the configured official ATB ZIP without network access.",
    )
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(message)s")
    args = parse_args()
    try:
        output_dir = fetch_and_normalize(
            args.scenario,
            download=not args.no_download,
        )
    except (FileNotFoundError, NlrAtbAutonomieError, requests.RequestException) as exc:
        raise SystemExit(f"NLR ATB/ANL adapter failed: {exc}") from exc
    logging.info("Wrote NLR ATB/ANL interim outputs to %s", output_dir)


if __name__ == "__main__":
    main()
