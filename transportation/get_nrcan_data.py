import os
import shutil
import pandas as pd
import openpyxl

spreadsheet_name = 'CANOE_TRN_<r>_v3'    # Copies will be made of the master_spreadsheet (ON) after updating with NRCan EUD tables from other provinces
province_list = [
    'QC', 
    'MB', 
    'SK', 
    'AB', 
    'BCT'
]   # Atlantic provinces are not compiled at the moment
nrcan_url = 'https://oee.nrcan.gc.ca/corporate/statistics/neud/dpa/data_e/downloads/comprehensive/Excel/2021/tran_<r>_e_<t>.xls'    # 2021 as latest year
nrcan_tables = {
    21: 'Car Explanatory', # car exp
    37: 'Truck Explanatory', # truck exp
    31: 'Bus Explanatory', # bus exp
    20: 'Cars', # car sec/act
    34: 'Passenger Light Trucks', # pass light truck sec/act
    35: 'Freight Light Trucks',   # freight light truck sec/act
    36: ['Medium Trucks', 'Heavy Trucks'], # mhd truck sec/act
    32: 'Motorcycle',  # motorcycle all
    28: 'School Buses', # school bus sec/act
    29: 'Urban Transit', # transit bus sec/act
    30: 'Inter-City Buses', # inter-city bus sec/act
    14: 'Passenger Air', # pass air sec
    15: 'Freight Air', # freight air sec
    17: 'Passenger Rail', # pass rail sec
    18: 'Freight Rail', # freight rail sec
    19: 'Marine Freight', # marine sec
    7:  'Off-Road'   # off-road sec
}

# Update these paths according to your files' locations
dir_path = os.path.dirname(os.path.abspath(__file__)) + '/'
this_dir = os.path.realpath(os.path.dirname(__file__)) + "/"

spreadsheet = dir_path + 'spreadsheet_database/' + spreadsheet_name + '.xlsx'
dir_tables = this_dir + 'nrcan_eud_tables/'

def get_nrcan_url(region, table_number):
    return str(nrcan_url).replace('<r>', region.lower()).replace('<t>', str(table_number))

def string_cleaner(string):
    return ''.join(letter for letter in string if letter in '- /()–|%' or letter.isalnum())

def string_letters(string):
    return ''.join(letter for letter in string_cleaner(string) if letter not in 'Â²¹')

def clean_index(df):
    df.index = [string_letters(idx) for idx in df.index]

def get_data(url, file_type=None, cache_file_type=None, name=None, **kwargs) -> pd.DataFrame | None:
    # Get the original file name
    if name == None: name = url.split("/")[-1].split("\\")[-1]
    if file_type == None: file_type = url.split(".")[-1]
    file_type = file_type.lower()

    if cache_file_type == None:
        if "xl" in file_type: cache_file_type = "csv"
        else: cache_file_type = file_type
    
    # If file type is different from new file type
    if name.split(".")[-1] != cache_file_type: name = os.path.splitext(name)[0] + "."+cache_file_type
    cache_file = dir_tables + name
    
    data = None
    if os.path.isfile(cache_file):
        # Get from existing local cache
        if cache_file_type == "csv": data = pd.read_csv(cache_file, index_col=0, dtype='unicode')
        print(f"Got {name} from local cache.")   
    else:
        print(f"Downloading {name} ...")
        try:
            # Download from url
            if file_type == "csv": data = pd.read_csv(url, **kwargs)
            elif "xl" in file_type: data = pd.read_excel(url, **kwargs)
            data.columns = [str(col).strip() for col in data.columns]
        except Exception as e:
            print(f"Failed to download {url}")
            print(e)

        # Try to cache downloaded file
        try:
            if not os.path.exists(dir_tables): os.mkdir(dir_tables)
            if cache_file_type == "csv": data.to_csv(cache_file)
            print(f"Cached {name}.")
        except Exception as e:
            print(f"Failed to cache {cache_file}.")
            print(e)
    return data

def get_nrcan_data(region, table_number, table_label, first_row=0, last_row=None) -> pd.DataFrame:
    # Get the requested table and discard excess rows, clean up table
    df = get_data(get_nrcan_url(region, table_number), skiprows=10)
    df = df.loc[first_row::] if last_row is None else df.loc[first_row:last_row]
    df = df.drop("Unnamed: 0", axis=1).set_index('Unnamed: 1')
    df.index.name = None
    
    # Process index to append variable names and drop empty rows
    header = None
    labels = []
    for idx, row in df.iterrows():
        if pd.isna(idx):
            header = None
            labels.append(idx)
        elif header is None:
            if pd.isna(row['2000']):    # First year column of every series
                header = idx
                labels.append(idx)
            else: labels.append(idx)
        else:
            labels.append(f"{header}|{idx}")
    df.index = labels
    df.dropna(inplace=True)
    clean_index(df)

    # Drop rows containing "Shares" and "GHG"
    df = df[~df.index.str.contains("Shares|GHG", na=False)]

    # Process index for entries with "Activity" and "Energy Intensity"
    new_index = []
    activity_count = 0
    intensity_count = 0
    for idx in df.index:
        # for tables with more than one vehicle class, e.g., table 36
        if isinstance(table_label, list) and ("Activity" in idx or "Energy Intensity" in idx or "Energy Use by Energy Source" in idx):
            if "Activity" in idx:
                label_to_use = table_label[activity_count % len(table_label)]
                activity_count += 1
            elif "Energy Intensity" in idx:
                label_to_use = table_label[intensity_count % len(table_label)]
                intensity_count += 1
            elif "Energy Use by Energy Source" in idx:
                label_to_use = table_label[0]   # only first label is assigned
            new_index.append(f"{label_to_use}|{idx}")
        elif isinstance(table_label, str) and ("Activity" in idx or "Energy Intensity" in idx or "Energy Use by Energy Source" in idx):
            new_index.append(f"{table_label}|{idx}")
        else:
            new_index.append(idx)
    df.index = new_index
    df.columns = [int(col) for col in df.columns]
    
    # Convert all data from strings to floats or nan
    df.replace('n.a.', pd.NA, inplace=True)
    df = df.astype(float, errors='ignore')
    return df

def concatenate_all_tables(region, nrcan_tables):
    all_tables = []
    for table_number, table_label in nrcan_tables.items():
        print(f"Processing Table {table_number} ({table_label})...")
        try:
            processed_df = get_nrcan_data(region, table_number, table_label)
            all_tables.append(processed_df)
        except Exception as e:
            print(f"Failed to process Table {table_number}: {e}")
    
    df = pd.concat(all_tables, axis=0)
    return df

def compile_spreadsheets(region, spreadsheet, nrcan_tables, start_row=3, end_row=118, insert_col=3):
    """
    New spreadsheets are built on top of the ON spreadsheet model and copied into new spreadsheets
    """
    nrcan_df = concatenate_all_tables(region.lower(), nrcan_tables)
    master_spreadsheet = spreadsheet.replace('<r>', 'ON')

    target_spreadsheet = spreadsheet.replace('<r>', region)
    shutil.copyfile(master_spreadsheet, target_spreadsheet)

    workbook = openpyxl.load_workbook(target_spreadsheet)
    sheet = workbook['Background Data']

    # Iterate through the rows and update time-series data
    for i, row in enumerate(sheet.iter_rows(min_row=start_row + 1, max_row=end_row, max_col=1)):
        spreadsheet_index = row[0].value

        if spreadsheet_index in nrcan_df.index:  # Match the index with nrcan_df
            time_series = nrcan_df.loc[spreadsheet_index].values
            time_series = [float(value) if pd.notnull(value) else None for value in time_series]
            # print(f"{spreadsheet_index}: {time_series}\n")

            # Determine the range of columns to update, starting from column C
            for col_offset, value in enumerate(time_series):
                sheet.cell(row=start_row + i + 1, column=insert_col + col_offset, value=value)

    workbook.save(target_spreadsheet)
    print(f'Sucessfully re-created {target_spreadsheet}.')

for province in province_list:
    compile_spreadsheets(province, spreadsheet, nrcan_tables)